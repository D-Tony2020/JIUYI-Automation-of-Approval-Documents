# -*- coding: utf-8 -*-
"""M2.1 无拐杖走骨架:伪真单 → 真实管线 → 承认书,对标 golden 打分(全6有效案)。

拐杖全拆:FAI ← drawing_extract(多模态,真); 文件→表 ← file_router(真);
         文件→料行 ← content_match(PDF内容,真)。
A2/B(BOM 成分 cell) ← golden mock(A2 人工填 / B 已单独验 97%, 此处测集成非测B)。
打分:FAI cell(真A1) + 各表 OLE 计数(真C) + COM 可开。039397 损坏golden 排除。
"""
import os, sys, io, glob, datetime
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.drawing_extract import extract as draw_extract
from hitl.file_router import route
from hitl.file_match import content_match, match_material
from hitl.build import build_upto
from hitl.ole_assemble import make_icon, embed_many, count_ole, verify_open
from hitl.material_table import compute_layout, DATA_TOP
from hitl.ole_placement import ev_col, part_top, mat_anchors_nocrutch
from hitl import sample_photo
from study.embed_structure import (grid_anchors, GRID, MATCERT_W, MATCERT_H,
                                    MATCERT_PART_TOPS, MATCERT_X0, MATCERT_DX)
from study.golden_parse import parse_golden
from study.case_data import to_inject_bom, _distinct_bins_per_sheet, extract_ole_labels

ROOT = os.path.dirname(os.path.abspath(__file__))
PSEUDO_ROOT = os.path.join(ROOT, "本单输入", "pseudo")
BLANK = os.path.join(ROOT, "模板", "承认书空白模板_通用.xlsx")
CASES_DIR = os.path.join(ROOT, "案例材料", "承认书", "参考用承诺书集")
OUTROOT = os.environ.get("M21_OUT") or os.path.join(ROOT, "产出留档", "M2.1-走骨架")
TODAY = datetime.date(2026, 6, 25)
SKIP = {"YY60039397"}            # 损坏 golden(嵌错图)

SHORT_KW = {"材质表": "材质成分", "材质证明": "材质证明", "部件承认": "部件",
            "UL": "UL", "信赖性": "信赖", "包装": "包装", "出货": "出货", "图纸": "图纸"}
# 单OLE表固定模板位(取自golden, 各案近恒定; 非per-case crutch, 同GRID性质)
DEFAULT_ANCHOR = {"图纸": (183, 266, 152, 42), "包装": (208, 271, 112, 60), "出货": (193, 261, 159, 42)}


def full_sheet(names, short):
    kw = SHORT_KW[short]
    for n in names:
        if kw in n or (short == "出货" and ("出貨" in n or "检验" in n)):
            return n
    return None


def identify(pdf, base, flat):
    """料 identity:内容驱动优先(可靠), 认不出(繁体/英文content)回退文件名匹配。"""
    mi = content_match(pdf, flat)
    return mi if mi is not None else match_material(base, flat)


def write_ole_labels(wb, ole_labels):
    """OLE 表 BOM 料/件名标签写回。标签文本=A2(BOM)内容(骨架取自golden=A2 mock); 字体/位置随之。"""
    for sheet, labs in ole_labels.items():
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        for r, c, txt, font, align in labs:
            try:
                cell = ws.cell(r, c)
                cell.value, cell.font, cell.alignment = txt, font, align
            except (AttributeError, TypeError):
                pass


def run_case(code, golden):
    pseudo = os.path.join(PSEUDO_ROOT, code)
    outdir = os.path.join(OUTROOT, code)
    os.makedirs(os.path.join(outdir, "icons"), exist_ok=True)
    cell = os.path.join(outdir, "_cell.xlsx")
    out = os.path.join(outdir, f"{code}_无拐杖承认书.xlsx")
    bom = to_inject_bom(parse_golden(golden))
    flat = [{"材质": m["材质"], "零件": p["零件"]} for p in bom for m in p["materials"]]

    draw_pdf = glob.glob(os.path.join(pseudo, "drawing", "*.pdf"))[0]
    dd = draw_extract(draw_pdf)                                   # A1 真
    data = {"drawing_meta": {"名称": dd["名称"] or "导线", "品号": dd["品号"], "版本": dd["版本"]},
            "product": {"材料名称": dd["名称"] or "导线", "填表日期": TODAY},
            "bom": bom, "dimensions": dd["dimensions"]}
    build_upto(BLANK, cell, data, upto=4, highlight=False)
    wb = openpyxl.load_workbook(cell)
    sample_photo.fill_sample_photo(wb[sample_photo.SHEET], sorted(glob.glob(os.path.join(pseudo, "photos", "*"))))
    write_ole_labels(wb, extract_ole_labels(golden))             # OLE表料/件名标签(A2 mock)
    wb.save(cell)
    names = wb.sheetnames

    # 路由 + 内容匹配 → 候选; 按 一材一证 / 一材每证据列一份 去重(收高召回超报)
    specs, mt_specs = [], []
    seen_mt, seen_cert = set(), set()
    for f in sorted(glob.glob(os.path.join(pseudo, "materials", "*.pdf"))):
        base = os.path.basename(f)
        for short in route(base):                                # C 真(可扇出)
            fs = full_sheet(names, short)
            if not fs:
                continue
            if short == "材质表":
                mi, col = identify(f, base, flat), ev_col(base)  # 内容驱动+文件名回退 真
                if mi is None or (mi, col) in seen_mt:           # 认不出料的不硬塞(避错行+重叠)→确认环②; 同(料,列)一份
                    continue
                seen_mt.add((mi, col))
                sp = {"sheet": fs, "pdf": f, "short": short, "mat_idx": mi, "col": col, "W": 56, "H": 42}
                mt_specs.append(sp); specs.append(sp)
            elif short == "材质证明":
                mi = identify(f, base, flat)
                if mi is None or mi in seen_cert:                # 一材一证; 认不出料的不硬塞→确认环②
                    continue
                seen_cert.add(mi)
                specs.append({"sheet": fs, "pdf": f, "short": short, "mat_idx": mi, "W": 56, "H": 42})
            else:
                specs.append({"sheet": fs, "pdf": f, "short": short, "W": 56, "H": 42})
    specs.append({"sheet": full_sheet(names, "图纸"), "pdf": draw_pdf, "short": "图纸", "W": 320, "H": 220})
    for s in specs:
        s["icon"] = make_icon(s["pdf"], os.path.join(outdir, "icons", os.path.basename(s["pdf"]) + ".png"))

    for s, (r, c) in zip(mt_specs, mat_anchors_nocrutch(bom, mt_specs)):
        s["row"], s["col"] = r, c
    by_short = defaultdict(list)
    for s in specs:
        if s["short"] != "材质表":
            by_short[s["short"]].append(s)
    # 材质证明书:内容驱动落位 — 每份证书落到它那个零件行(按 content_match 的料→零件), 行内横排
    part_of_flat = [pi for pi, p in enumerate(bom) for _ in p["materials"]]
    cert_cnt = defaultdict(int)
    for s in by_short["材质证明"]:
        mi = s.get("mat_idx")
        pi = part_of_flat[mi] if (mi is not None and mi < len(part_of_flat)) else 0
        top = part_top(bom[pi]["零件"] if pi < len(bom) else "")   # 按零件类型对齐标签行
        s["L"], s["T"], s["W"], s["H"] = MATCERT_X0 + cert_cnt[top] * MATCERT_DX, top, MATCERT_W, MATCERT_H
        cert_cnt[top] += 1
    for short, gk in (("部件承认", "部件承认书"), ("UL", "UL证明"), ("信赖性", "信赖性")):
        g = GRID[gk]
        for s, (L, T) in zip(by_short[short], grid_anchors(len(by_short[short]), **g)):
            s["L"], s["T"], s["W"], s["H"] = L, T, g["w"], g["h"]
    for short in ("包装", "出货", "图纸"):
        L, T, W, H = DEFAULT_ANCHOR[short]
        for s in by_short[short]:
            s["L"], s["T"], s["W"], s["H"] = L, T, W, H
    for s in specs:
        if s["short"] != "材质表":
            s.setdefault("L", 95); s.setdefault("T", 250)
        s.setdefault("W", 56); s.setdefault("H", 42)

    embed_many(cell, out, specs)

    # 打分
    gwb = openpyxl.load_workbook(golden, data_only=True)
    fwb = openpyxl.load_workbook(out, data_only=True)
    gf = [s for s in gwb.sheetnames if "FAI" in s][0]
    ff = [s for s in fwb.sheetnames if "FAI" in s][0]
    fai_ok = fai_tot = 0
    ndim = len(dd["dimensions"])
    for r in range(9, 9 + ndim):
        for c in (2, 3, 4):
            fai_tot += 1
            fai_ok += 1 if str(gwb[gf].cell(r, c).value) == str(fwb[ff].cell(r, c).value) else 0
    go = {k: len(v) for k, v in _distinct_bins_per_sheet(golden).items()}
    fo = {k: len(v) for k, v in _distinct_bins_per_sheet(out).items()}
    exact = sum(1 for k in go if go[k] == fo.get(k, 0))
    over = {k.split(".")[-1][:6]: fo[k] - go.get(k, 0) for k in fo if fo[k] > go.get(k, 0)}
    try:
        opens = verify_open(out)["_total"]
    except Exception:
        opens = -1
    return {"code": code, "fai": (fai_ok, fai_tot), "sheet_exact": (exact, len(go)),
            "over": over, "ole": count_ole(out), "opens": opens}


def main():
    import re
    os.makedirs(OUTROOT, exist_ok=True)
    only = sys.argv[1] if len(sys.argv) > 1 else None     # 可选:只跑某案(传案号片段)
    cases = sorted([p for p in glob.glob(os.path.join(CASES_DIR, "*.xlsx"))
                    if not os.path.basename(p).startswith("~$")])
    print(f"=== M2.1 无拐杖走骨架 | {'单案 '+only if only else '全有效案'} ===\n")
    for g in cases:
        code = re.sub(r"[^A-Za-z0-9]", "", os.path.basename(g))[:10]
        if only and only not in code:
            continue
        if code in SKIP:
            print(f"⏭️  {code}: 跳过(损坏golden)"); continue
        try:
            r = run_case(code, g)
        except Exception as e:
            print(f"❌ {code}: {str(e)[:90]}"); continue
        ov = " ".join(f"{k}+{v}" for k, v in r["over"].items()) or "无"
        flag = "✅" if r["fai"][0] == r["fai"][1] and r["opens"] == r["ole"] else "⚠️"
        print(f"{flag} {code}: FAI {r['fai'][0]}/{r['fai'][1]} | 表精确 {r['sheet_exact'][0]}/{r['sheet_exact'][1]} "
              f"| 超报[{ov}] | OLE {r['ole']} 开{r['opens']}")
    print("\n超报=MSDS高召回(材质表/材质证明)→确认环②裁; 其余表精确; 全程零golden拐杖")


if __name__ == "__main__":
    main()
