# -*- coding: utf-8 -*-
"""W4 材质成分表 · 最小执行线 = "锡"一行(spike mock)。

build_upto(3)=封面+送样目录+材质表锡行；空白参照=build_upto(2)(材质表空)。
自检：锡行关键值(尤其 M14=63 非ND陷阱、CAS去空格、wt%→小数、<3原样) + 嵌套合并 + 底部表浮动。
"""
import os
import sys
import io
import json
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.build import build_upto
from hitl.harness import assert_no_external_links
from hitl import material_table

ROOT = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(ROOT, "模板", "承认书空白模板_治病.xlsx")
MOCK = os.path.join(ROOT, "spike", "mock")
OUTDIR = os.path.join(ROOT, "产出留档", "W4-材质成分表")
BLANK = os.path.join(OUTDIR, "材质表_空白参照__00.xlsx")
OUT = os.path.join(OUTDIR, "材质表_锡一行__01.xlsx")
SHEET = material_table.MAT_SHEET

# 锡 BOM ← spike 冻结 mock(签认) → assemble_row 归一
_msds = json.load(open(os.path.join(MOCK, "msds_extract.json"), encoding="utf-8"))
_rohs = json.load(open(os.path.join(MOCK, "rohs_extract.json"), encoding="utf-8"))
_row = material_table.assemble_row(_msds, _rohs, "锡", "锡丝", "锡")
BOM = [{"零件": _row["零件"], "供应商": _row["供应商"], "materials": [
    {"材质类别": _row["材质类别"], "材质": _row["材质"], "blocks": [
        {"报告编号": _row["检测报告编号"], "报告日期": _row["检测报告日期"],
         "RoHS": _row["RoHS"], "成份": _row["成份"]}]}]}]

DATA = {
    "drawing_meta": {"名称": "SB120420BLCNR009导线", "品号": "YY60039403", "版本": "A01"},
    "product": {"材料名称": "导线", "填表日期": datetime.date(2026, 6, 24)},
    "bom": BOM,
}


def main():
    build_upto(TPL, BLANK, DATA, upto=2)  # 材质表空
    build_upto(TPL, OUT, DATA, upto=3)    # + 材质表锡行

    ws = openpyxl.load_workbook(OUT)[SHEET]
    checks = {
        "A14": 1, "B14": "锡", "C14": "兴鸿泰", "D14": "锡丝", "E14": "锡",
        "G14": "SN", "H14": "7440-31-5", "J14": "0.993",
        "G15": "CU", "H15": "7440-50-8", "J15": "0.007",
        "G16": "改性松香", "H16": "65997-05-9", "J16": "<3",
        "M14": "63", "N14": "ND", "W14": "2025.06.30", "X14": "SZXEC25002243403",
        "A2": "供货商类别  :导线",
    }
    errs = material_table.selfcheck_material(ws, checks)
    assert_no_external_links(OUT)

    # 嵌套合并：材质块 D14:D16 / 项次块 A14:A16
    merges = {str(m) for m in ws.merged_cells.ranges}
    for need in ("D14:D16", "A14:A16", "M14:M16"):
        if need not in merges:
            errs.append(f"缺合并 {need}")
    # J2 不再是模板硬编码 2026-06-18
    if "2026 年   06   月  18" in str(ws["J2"].value):
        errs.append("J2 仍是模板硬编码日期")
    # 底部表浮动：数据末行16 → 底部表起于 19
    if not (ws["A19"].value and "RoHS排除管制对象" in str(ws["A19"].value)):
        errs.append(f"底部表未浮动到 A19，实得 {ws['A19'].value!r}")

    status = "PASS" if not errs else "FAIL"
    os.makedirs(OUTDIR, exist_ok=True)
    with open(os.path.join(OUTDIR, "_manifest.txt"), "a", encoding="utf-8") as f:
        f.write(f"材质表_锡一行__01\tPb=M14={ws['M14'].value} 底部表@A19\t{status}\n")

    print("空白参照:", BLANK)
    print("填入成品:", OUT)
    print("锡行关键值: A14=%s B14=%s C14=%s | H14=%s J14=%s | M14(Pb)=%s W14=%s X14=%s" % (
        ws["A14"].value, ws["B14"].value, ws["C14"].value, ws["H14"].value, ws["J14"].value,
        ws["M14"].value, ws["W14"].value, ws["X14"].value))
    print("J2 填表日期:", ws["J2"].value)
    print("底部表起始 A19:", ws["A19"].value)
    if errs:
        print("❌ 自测失败：")
        for e in errs:
            print("   -", e)
        sys.exit(1)
    print("✅ 自测通过：锡行值正确(Pb=63非ND/CAS去空格/wt%→小数/<3原样) + 嵌套合并 + J2修复 + 底部表浮动")


if __name__ == "__main__":
    main()
