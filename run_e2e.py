# -*- coding: utf-8 -*-
"""7 案端到端全量测试：通用空白模板 + 各案本单输入(自身内嵌文件) → 封装承认书，对标 golden。

产出(产出留档/E2E/)供手测：
  通用空白模板.xlsx        — 纯静态、可复用
  <案>_图纸特异性空白.xlsx — 结构搭好(材质表行/零件材质/材质证明标签/FAI行)、动态数据空
  <案>_封装承认书.xlsx     — cells + 材质证明标签 + 全OLE(golden位置忠实) + 照片
对标 golden：材质表/封面/FAI 数据 + 各表 OLE 计数 + 照片数。
"""
import os
import sys
import io
import shutil
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.build import build_upto
from hitl.ole_assemble import make_icon, embed_many, count_ole
from hitl import sample_photo
from study.case_data import extract_case, to_inject_bom, extract_drawing_meta, extract_dimensions
from study.golden_parse import parse_golden
from study.ole_structure import ole_count_per_sheet

ROOT = os.path.dirname(os.path.abspath(__file__))
BLANK = os.path.join(ROOT, "模板", "承认书空白模板_通用.xlsx")
CASES_DIR = os.path.join(ROOT, "案例材料", "承认书", "参考用承诺书集")
SRC_DIR = os.path.join(ROOT, "本单输入", "test_cases")
OUTDIR = os.path.join(ROOT, "产出留档", "E2E")
MAT = "材质成分展开表"
PHOTO = sample_photo.SHEET
# 材质表数据列(scaffold 清空, 留结构 A-F)
MAT_DATA_COLS = [7, 8, 9, 10] + list(range(13, 25)) + [26, 27, 28, 29, 30]


def _n(v):
    if v is None:
        return ""
    return ("%g" % v) if isinstance(v, float) else str(v).strip()


def _msheet(wb):
    return wb[[s for s in wb.sheetnames if MAT in s][0]]


def _set(ws, r, c, v):
    """安全写格(合并格只读则跳)。"""
    try:
        ws.cell(r, c).value = v
    except (AttributeError, TypeError):
        pass


def write_matcert_labels(wb, labels):
    ws = [wb[s] for s in wb.sheetnames if "材质证明" in s][0]
    for r, txt in labels:
        _set(ws, r, 2, txt)


def build_scaffold(cell_xlsx, scaffold_xlsx, labels):
    """图纸特异性空白：结构(行/零件/材质/标签/FAI行)在, 动态数据清空。"""
    wb = openpyxl.load_workbook(cell_xlsx)
    mat = _msheet(wb)
    for r in range(14, mat.max_row + 1):
        for c in MAT_DATA_COLS:
            _set(mat, r, c, None)
    cov = wb["封面"]
    for r, c in ((12, 4), (14, 4), (16, 4)):
        _set(cov, r, c, None)
    fws = [wb[s] for s in wb.sheetnames if "FAI" in s]
    if fws:
        for r in range(9, 40):
            for c in (2, 3, 4):
                _set(fws[0], r, c, None)
    write_matcert_labels(wb, labels)
    wb.save(scaffold_xlsx)


def run_case(golden):
    cd = extract_case(golden, SRC_DIR)
    code = cd["code"]
    data = {
        "drawing_meta": {"名称": cd["drawing_meta"]["品类"], "品号": cd["drawing_meta"]["品号"],
                         "版本": cd["drawing_meta"]["版本"]},
        "product": {"材料名称": cd["drawing_meta"]["品类"], "填表日期": datetime.date(2026, 6, 25)},
        "bom": cd["bom"],
        "dimensions": cd["dimensions"],
    }
    # 每案三位一体: 一个文件夹含 1通用/2图纸特异性/3封装
    folder = os.path.join(OUTDIR, code)
    os.makedirs(folder, exist_ok=True)
    shutil.copy(BLANK, os.path.join(folder, "1_通用空白模板.xlsx"))
    cell = os.path.join(folder, "_cell.xlsx")
    scaffold = os.path.join(folder, "2_图纸特异性空白模板.xlsx")
    final = os.path.join(folder, "3_封装承认书.xlsx")

    # 段一: cells + 材质证明标签 + 照片
    build_upto(BLANK, cell, data, upto=4, highlight=False)
    build_scaffold(cell, scaffold, cd["matcert_labels"])     # 图纸特异性空白(数据清前先存结构)
    wb = openpyxl.load_workbook(cell)
    write_matcert_labels(wb, cd["matcert_labels"])
    sample_photo.fill_sample_photo(wb[PHOTO], cd["photos"])
    wb.save(cell)

    # 段二: 装全 OLE(golden 位置忠实复现)
    specs = []
    for s in cd["ole_map"]:
        icon = s["pdf"] + ".png"
        make_icon(s["pdf"], icon)
        specs.append({"sheet": s["sheet"], "pdf": s["pdf"], "icon": icon,
                      "L": s["L"], "T": s["T"], "W": s["W"], "H": s["H"]})
    embed_many(cell, final, specs)

    # 对标 golden
    gwb = openpyxl.load_workbook(golden, data_only=True)
    fwb = openpyxl.load_workbook(final, data_only=True)
    gmat, fmat = _msheet(gwb), _msheet(fwb)
    mats = parse_golden(golden)
    last = max((c["_row"] for m in mats for c in m["components"]), default=13)
    data_diff = 0
    for r in range(14, last + 1):
        for c in [1, 2, 3, 5, 7, 8, 10] + list(range(13, 25)) + [28, 29]:  # 不含材质类别col4
            if _n(gmat.cell(r, c).value) != _n(fmat.cell(r, c).value):
                data_diff += 1
    g_ole, f_ole = ole_count_per_sheet(golden), ole_count_per_sheet(final)
    ole_match = sum(1 for k in g_ole if g_ole[k] == f_ole.get(k, 0))
    g_photo = len(cd["photos"])
    f_photo = sum(1 for im in fwb[[s for s in fwb.sheetnames if s.strip() == PHOTO.strip()][0]]._images
                  if sample_photo._anchor_col(im) != 0)
    if os.path.exists(cell):
        os.remove(cell)      # 删中间产物
    return {"code": code, "材质": len(mats), "尺寸": len(cd["dimensions"]),
            "ole": count_ole(final), "data_diff": data_diff,
            "ole_sheet_match": (ole_match, len(g_ole)), "photo": (f_photo, g_photo),
            "final": final, "scaffold": scaffold}


def main():
    os.makedirs(OUTDIR, exist_ok=True)
    cases = sorted([p for p in __import__("glob").glob(os.path.join(CASES_DIR, "*.xlsx"))
                    if not os.path.basename(p).startswith("~$")])
    print(f"=== 7 案端到端全量测试 ===\n")
    for g in cases:
        try:
            r = run_case(g)
        except Exception as e:
            print(f"{os.path.basename(g)[:10]}: ❌ {str(e)[:120]}")
            continue
        flag = "✅" if (r["data_diff"] == 0 and r["ole_sheet_match"][0] == r["ole_sheet_match"][1]
                       and r["photo"][0] == r["photo"][1]) else "⚠️"
        print(f"{flag} {r['code']}: 材质{r['材质']} 尺寸{r['尺寸']} | OLE装{r['ole']} "
              f"各表计数{r['ole_sheet_match'][0]}/{r['ole_sheet_match'][1]} "
              f"照片{r['photo'][0]}/{r['photo'][1]} 数据diff{r['data_diff']}")
    print(f"\n产出(产出留档/E2E/<案>/): 每案三位一体 = 1_通用空白模板 + 2_图纸特异性空白模板 + 3_封装承认书")


if __name__ == "__main__":
    main()
