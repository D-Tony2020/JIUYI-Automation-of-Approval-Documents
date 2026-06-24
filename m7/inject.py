# -*- coding: utf-8 -*-
"""M7 注入引擎（填入）：把封面元数据 + 材质成分表变行 注入黄金模板。

材质表列(1-indexed)：A项次 B零件 C供应商 D材质类别 E材质 F材质重量 G成份名 H CAS
 I weight(g) J weight% K MSDS L Reach M..V RoHS十项 W报告日期 X报告编号 Y RoHs2.0 Z Br AA Cl
 AB是否符合 AC是否RoHS排外 AD排除项次
嵌套合并：A/B/C 跨整个"零件"块；D/E/F/K/L/M..V/W/X/Y/Z/AA/AB/AC/AD 跨同一"材质"的成份行；
G/H/I/J 每个成份一行(不合并)。
"""
import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from copy import copy

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
FONT = Font(name="宋体", size=9)

ROHS_ORDER = ["Pb", "Cd", "Hg", "Cr6+", "PBBs", "PBDEs", "DEHP", "DBP", "BBP", "DIBP"]
ROHS_COL = {k: 13 + i for i, k in enumerate(ROHS_ORDER)}   # M..V = 13..22
MAT_MERGE_COLS = [4, 5, 6, 11, 12] + list(range(13, 31))    # D,E,F,K,L,M..AD
PART_MERGE_COLS = [1, 2, 3]                                  # A,B,C


def _style_cell(ws, r, c):
    cell = ws.cell(r, c)
    cell.border = BORDER
    cell.alignment = CENTER
    cell.font = FONT


def _clear_data_region(ws, top=14, bottom=51):
    """拆掉数据区残留的旧合并并清值(保留表头≤13与底部附表≥52)。"""
    for mr in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = mr.bounds
        if min_r >= top and max_r <= bottom:
            ws.unmerge_cells(str(mr))
    for row in ws.iter_rows(min_row=top, max_row=bottom, min_col=1, max_col=30):
        for cell in row:
            cell.value = None


def inject_material_table(ws, bom, start_row=14):
    _clear_data_region(ws)
    r = start_row
    for idx, part in enumerate(bom, 1):
        part_start = r
        for mat in part["materials"]:
            mat_start = r
            for comp in mat["成份"]:
                ws.cell(r, 7, comp.get("成份名称", ""))
                ws.cell(r, 8, comp.get("CAS", ""))
                ws.cell(r, 10, comp.get("重量%", ""))
                r += 1
            mat_end = r - 1
            ws.cell(mat_start, 4, mat.get("材质类别", ""))
            ws.cell(mat_start, 5, mat.get("材质", ""))
            for k in ROHS_ORDER:
                ws.cell(mat_start, ROHS_COL[k], mat.get("RoHS", {}).get(k, ""))
            ws.cell(mat_start, 23, mat.get("报告日期", ""))
            ws.cell(mat_start, 24, mat.get("报告编号", ""))
            ws.cell(mat_start, 28, "Yes")   # 是否符合
            ws.cell(mat_start, 29, "否")     # 是否RoHS排外
            if mat_end > mat_start:
                for c in MAT_MERGE_COLS:
                    ws.merge_cells(start_row=mat_start, start_column=c, end_row=mat_end, end_column=c)
        part_end = r - 1
        ws.cell(part_start, 1, idx)
        ws.cell(part_start, 2, part.get("零件", ""))
        ws.cell(part_start, 3, part.get("供应商", ""))
        if part_end > part_start:
            for c in PART_MERGE_COLS:
                ws.merge_cells(start_row=part_start, start_column=c, end_row=part_end, end_column=c)
        # 样式
        for rr in range(part_start, part_end + 1):
            for cc in range(1, 31):
                _style_cell(ws, rr, cc)
    return r - 1  # 最后写入行


def inject(template_path, out_path, product, bom):
    wb = openpyxl.load_workbook(template_path)   # data_only=False 保留公式
    # 封面
    cov = wb["封面"]
    cov["D12"] = product["材料名称"]
    cov["D14"] = product["料号"]
    cov["D16"] = product["版本"]
    # 材质成分表
    mz = next(s for s in wb.sheetnames if "材质成分展开" in s)
    ws = wb[mz]
    ws["A2"] = f"供货商类别  :{product['材料名称']}"
    ws["A3"] = f"产品名称    :{product['料号']}"
    last = inject_material_table(ws, bom)
    wb.save(out_path)
    return last


if __name__ == "__main__":
    import os, sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    ROOT = r"D:\Desktop\Moore 工业智能\久益\久益-承认书自动化"
    TPL = os.path.join(ROOT, r"模板\承认书空白模板.xlsx")
    OUT = os.path.join(ROOT, r"模板\填充示例_锡+热缩管.xlsx")

    product = {"材料名称": "导线", "料号": "YY60039403-DEMO", "版本": "A01"}
    ND9 = {k: "ND" for k in ROHS_ORDER}
    bom = [
        {"零件": "热缩管", "供应商": "领飞", "materials": [
            {"材质类别": "套管", "材质": "聚烯烃", "报告编号": "A225086578510100101E", "报告日期": "2025.11.24",
             "RoHS": dict(ND9),
             "成份": [
                 {"成份名称": "EVA", "CAS": "24937-78-8", "重量%": "0.5"},
                 {"成份名称": "氢氧化镁", "CAS": "1309-42-8", "重量%": "0.4"},
                 {"成份名称": "氮系阻燃剂", "CAS": "37640-57-6", "重量%": "0.06"},
                 {"成份名称": "色母", "CAS": "/", "重量%": "0.04"},
             ]},
        ]},
        {"零件": "锡", "供应商": "兴鸿泰", "materials": [
            {"材质类别": "锡丝", "材质": "锡", "报告编号": "SZXEC25002243403", "报告日期": "2025.06.30",
             "RoHS": dict(ND9, Pb="63"),
             "成份": [
                 {"成份名称": "SN", "CAS": "7440-31-5", "重量%": "0.993"},
                 {"成份名称": "CU", "CAS": "7440-50-8", "重量%": "0.007"},
                 {"成份名称": "改性松香", "CAS": "65997-05-9", "重量%": "<3"},
             ]},
        ]},
    ]
    last = inject(TPL, OUT, product, bom)
    print(f"✅ 注入完成，材质表写到第 {last} 行 -> {OUT}")
