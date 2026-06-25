# -*- coding: utf-8 -*-
"""每案数据抽取器：从一份做好的承认书(golden) 抽出"本单输入" —— 模拟用户上传的结构输入。

产出: drawing_meta(品号/品类/版本) + dimensions(FAI尺寸) + bom(材质表) + parts(零件) +
      ole_map(每表内嵌文件) + photos(样品照片)。供 case-agnostic 生成器回放, 再对标 golden。
"""
import os
import re
import sys
import zipfile

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
import openpyxl
from study.golden_parse import parse_golden, find_mat_sheet
from hitl.ole_assemble import com_session, extract_embedded_pdf, original_filename
from hitl.sample_photo import _anchor_col

PHOTO_SHEET = "4.样品照片（多角度）"
FAI_KW = "FAI"


def _sheet(wb, kw):
    for s in wb.sheetnames:
        if kw in s:
            return wb[s]
    return None


def extract_drawing_meta(golden):
    wb = openpyxl.load_workbook(golden, data_only=True)
    ws = [wb[s] for s in wb.sheetnames if s.strip() == "封面"][0]
    return {"品号": str(ws["D14"].value or "").strip(),
            "品类": str(ws["D12"].value or "").strip(),
            "版本": str(ws["D16"].value or "").strip()}


def extract_dimensions(golden):
    """FAI 规格 B/C/D = LSL/中心/USL → (标称=中心, 上=USL-中心, 下=中心-LSL)。"""
    wb = openpyxl.load_workbook(golden, data_only=True)
    ws = _sheet(wb, FAI_KW)
    dims = []
    if ws is None:
        return dims
    for r in range(9, 45):
        b, c, d = ws.cell(r, 2).value, ws.cell(r, 3).value, ws.cell(r, 4).value
        try:
            lsl, ctr, usl = float(b), float(c), float(d)
        except (TypeError, ValueError):
            continue
        dims.append((ctr, round(usl - ctr, 4), round(ctr - lsl, 4)))
    return dims


def _distinct_bins_per_sheet(xlsx):
    z = zipfile.ZipFile(xlsx)
    names = set(z.namelist())
    wbx = z.read("xl/workbook.xml").decode("utf-8", "ignore")
    sheets = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbx)
    rel = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"',
                          z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")))
    out = {}
    for nm, rid in sheets:
        wsf = rel[rid]
        wsf = ("xl/" + wsf) if not wsf.startswith("xl/") else wsf
        base = os.path.basename(wsf)
        relf = f"xl/worksheets/_rels/{base}.rels"
        if relf not in names:
            continue
        tg = re.findall(r'Target="([^"]+)"', z.read(relf).decode("utf-8", "ignore"))
        bins = sorted({os.path.basename(t) for t in tg if "embeddings/oleObject" in t},
                      key=lambda b: int(re.findall(r"\d+", b)[0]))
        if bins:
            out[nm.replace("&amp;", "&")] = bins
    return out


def extract_ole_map(golden, out_dir):
    """每表内嵌 OLE: distinct bins + 抽 PDF + golden 位置(COM, 供 spatial_order 对应)。"""
    os.makedirs(out_dir, exist_ok=True)
    sheet_bins = _distinct_bins_per_sheet(golden)
    code = re.sub(r"[^A-Za-z0-9]", "", os.path.basename(golden))[:12]
    manifest = []
    _used = set()
    with com_session() as xl:
        wb = xl.Workbooks.Open(os.path.abspath(golden), UpdateLinks=0, ReadOnly=True)
        for sh in wb.Worksheets:
            bins = sheet_bins.get(sh.Name) or sheet_bins.get(sh.Name.strip())
            if not bins:
                continue
            n = sh.OLEObjects().Count
            geos = []
            for i in range(1, n + 1):
                o = sh.OLEObjects().Item(i)
                try:
                    geos.append((o.Left, o.Top, o.Width, o.Height, o.TopLeftCell.Row, o.TopLeftCell.Column))
                except Exception:
                    geos.append((0, 0, 56, 42, 1, 1))
            for k, b in enumerate(bins):
                g = geos[k] if k < len(geos) else (geos[-1] if geos else (0, 0, 56, 42, 1, 1))
                orig = original_filename(golden, b) or b.replace(".bin", ".pdf")   # 用原始文件名
                fname = re.sub(r'[\\/:*?"<>|]', "_", orig)
                if not fname.lower().endswith(".pdf"):
                    fname += ".pdf"
                pdf = os.path.join(out_dir, fname)
                dup = 1
                while pdf in _used:                       # 同名去重
                    pdf = os.path.join(out_dir, fname[:-4] + f"_{dup}.pdf")
                    dup += 1
                _used.add(pdf)
                try:
                    extract_embedded_pdf(golden, b, pdf)
                except Exception:
                    continue
                manifest.append({"sheet": sh.Name, "bin": b, "pdf": pdf,
                                 "L": round(g[0], 1), "T": round(g[1], 1),
                                 "W": round(g[2], 1), "H": round(g[3], 1),
                                 "row": g[4], "col": g[5]})
        wb.Close(False)
    return manifest


_LABEL_HEADER = ("生久", "http", "承认书", "证明书", "测试报告", "报告", "适用",
                 "REACH", "MSDS", "Reach")


def extract_ole_labels(golden):
    """所有 OLE 嵌入组表的 BOM 物料标签 {sheet名: [(row, col, text, font, align)]}。

    连同 golden 的字体/对齐一起捕获, 写回时还原 → 字体统一 + 与图标对齐(与原版一致)。
    """
    import copy
    wb = openpyxl.load_workbook(golden, data_only=True)
    out = {}
    for s in wb.sheetnames:
        if not any(k in s for k in ("部件", "信赖", "材质证明", "UL", "ul")):
            continue
        ws = wb[s]
        labs = []
        for r in range(11, 26):
            for c in (2, 3):
                cell = ws.cell(r, c)
                v = cell.value
                if v and not any(k in str(v) for k in _LABEL_HEADER):
                    labs.append((r, c, str(v).strip(),
                                 copy.copy(cell.font), copy.copy(cell.alignment)))
        if labs:
            out[s] = labs
    return out


def extract_photos(golden, out_dir):
    os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.load_workbook(golden)
    ws = [wb[s] for s in wb.sheetnames if s.strip() == PHOTO_SHEET.strip()]
    if not ws:
        return []
    paths = []
    for i, im in enumerate(ws[0]._images):
        if _anchor_col(im) == 0:
            continue
        p = os.path.join(out_dir, f"photo_{i}.png")
        with open(p, "wb") as f:
            f.write(im._data())
        paths.append(p)
    return paths


def to_inject_bom(materials):
    """parse_golden 的扁平 materials → inject_data 需要的嵌套 bom(零件→材质→报告块→成份)。

    成份按 _row 落到所属报告块(块区间=本块_first 到下块_first-1)。供 fill_material_table 回放。
    """
    parts = {}
    order = []
    for m in materials:
        p = (m["零件"] or "").strip()
        if p not in parts:
            parts[p] = {"零件": p, "供应商": (m.get("供应商") or "").strip(),
                        "材质类别": (m.get("材质类别") or "").strip(), "materials": []}
            order.append(p)
        first_of_part = len(parts[p]["materials"]) == 0   # 材质类别零件级合并: 只首材质写, 余空
        blks = sorted(m["blocks"], key=lambda b: b["_first"])
        comps = sorted(m["components"], key=lambda c: c["_row"])
        out_blocks = []
        for i, blk in enumerate(blks):
            lo = blk["_first"]
            hi = blks[i + 1]["_first"] - 1 if i + 1 < len(blks) else 10 ** 9
            bc = [c for c in comps if lo <= c["_row"] <= hi]
            out_blocks.append({
                "成份": [{"成份名称": c["成份名"], "CAS": c["CAS"], "重量%": c["重量%"]} for c in bc],
                "RoHS": blk.get("rohs", {}),
                "报告编号": blk.get("报告号", ""),
                "报告日期": blk.get("报告日期", ""),
            })
        parts[p]["materials"].append({"材质类别": (m["材质类别"] if first_of_part else ""),
                                       "材质": m["材质"], "blocks": out_blocks})
    return [parts[p] for p in order]


def extract_case(golden, root_out):
    code = re.sub(r"[^A-Za-z0-9]", "", os.path.basename(golden))[:10]
    d = os.path.join(root_out, code)
    bom_mats = parse_golden(golden)
    bom = to_inject_bom(bom_mats)          # 嵌套块格式(供 inject_data/compute_layout/count_from_bom)
    order = [p["零件"] for p in bom]
    return {
        "code": code,
        "golden": golden,
        "drawing_meta": extract_drawing_meta(golden),
        "dimensions": extract_dimensions(golden),
        "bom_materials": bom_mats,
        "bom": bom,
        "parts": order,
        "ole_map": extract_ole_map(golden, os.path.join(d, "ole")),
        "photos": extract_photos(golden, os.path.join(d, "photos")),
        "ole_labels": extract_ole_labels(golden),
    }


if __name__ == "__main__":
    import io
    import json
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    G = sys.argv[1] if len(sys.argv) > 1 else \
        r"案例材料\承认书\参考用承诺书集\YY60010122 (J00010327）承认书.xlsx"
    c = extract_case(G, r"本单输入\test_cases")
    print("案例:", c["code"])
    print("  图纸:", c["drawing_meta"])
    print("  尺寸:", c["dimensions"])
    print("  零件:", c["parts"], " 材质数:", len(c["bom_materials"]))
    print("  OLE:", len(c["ole_map"]), "个  照片:", len(c["photos"]), "张")
