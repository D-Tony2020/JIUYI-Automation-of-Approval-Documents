# -*- coding: utf-8 -*-
"""治病模板 → 通用空白模板：清掉所有 YY60039403 专属动态内容，只留确定的静态组件。

清：① FAI 品类/日期/规格(B-D)/实测(I-AN)；② 所有 OLE 嵌入组表的 BOM 物料文本标签
    (部件承认书/信赖性/材质证明书/UL 的 线材/端子… col B/C)；③ 样品照片残照(留LOGO)。
留：公司抬头、表头、FAI 公式(E-H)、花名册、页脚等静态件。
"""
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
import openpyxl

SRC = "模板/承认书空白模板_治病.xlsx"
DST = "模板/承认书空白模板_通用.xlsx"
OLE_LABEL_SHEETS = ("部件", "信赖", "材质证明", "UL", "ul")
HEADER_KW = ("生久", "http", "承认书", "证明书", "测试报告", "报告", "适用", "REACH", "MSDS", "Reach")


def _set(ws, r, c, v):
    try:
        ws.cell(r, c).value = v
    except (AttributeError, TypeError):
        pass


def build(root):
    wb = openpyxl.load_workbook(os.path.join(root, SRC))
    cleared = {"FAI": 0, "标签": [], "照片": 0}

    # ① FAI 清动态(留 E-H 公式 col5-8 + 花名册)
    fws = [wb[s] for s in wb.sheetnames if "FAI" in s]
    if fws:
        f = fws[0]
        _set(f, 2, 3, None)   # C2 品类
        _set(f, 2, 21, None)  # U2 日期
        for r in range(9, 44):
            for c in [2, 3, 4] + list(range(9, 41)):   # B/C/D 规格 + I-AN 实测
                if f.cell(r, c).value not in (None, ""):
                    _set(f, r, c, None)
                    cleared["FAI"] += 1

    # ② OLE 表 BOM 物料标签
    for s in wb.sheetnames:
        if not any(k in s for k in OLE_LABEL_SHEETS):
            continue
        sh = wb[s]
        for r in range(11, 26):
            for c in (2, 3):
                v = sh.cell(r, c).value
                if v and not any(k in str(v) for k in HEADER_KW):
                    cleared["标签"].append((s[:8], sh.cell(r, c).coordinate, str(v).strip()[:8]))
                    _set(sh, r, c, None)

    # ③ 样品照片残照(留 LOGO at col0)
    from hitl.sample_photo import _anchor_col, SHEET as PSHEET
    pw = [wb[s] for s in wb.sheetnames if s.strip() == PSHEET.strip()]
    if pw:
        n0 = len(pw[0]._images)
        pw[0]._images = [im for im in pw[0]._images if _anchor_col(im) == 0]
        cleared["照片"] = n0 - len(pw[0]._images)

    wb.save(os.path.join(root, DST))
    # ④ 公章去背景(zip手术换 media: 白底→透明, 避 openpyxl 图片流问题)
    cleared["公章"] = seal_transparent_zip(os.path.join(root, DST))
    return cleared


def seal_transparent_zip(xlsx):
    """zip 手术: 公章(670×567)白底转透明留红章; JPEG→PNG 改名 + 更新 rels/content-types。"""
    import io as _io
    import os as _os
    import zipfile
    from PIL import Image as PI
    z = zipfile.ZipFile(xlsx)
    names = z.namelist()
    target = None
    for n in names:
        if not n.startswith("xl/media/"):
            continue
        try:
            if PI.open(_io.BytesIO(z.read(n))).size == (670, 567):
                target = n
                break
        except Exception:
            pass
    if not target:
        z.close()
        return 0
    p = PI.open(_io.BytesIO(z.read(target))).convert("RGBA")
    px = [(r, g, bl, 0) if (r > 230 and g > 230 and bl > 230) else (r, g, bl, a)
          for r, g, bl, a in list(p.getdata())]
    p.putdata(px)
    buf = _io.BytesIO()
    p.save(buf, format="PNG")
    new_png = buf.getvalue()
    old_base = _os.path.basename(target)
    new_base = old_base.rsplit(".", 1)[0] + ".png"     # image4.jpeg → image4.png
    data_map = {n: z.read(n) for n in names}
    z.close()
    with zipfile.ZipFile(xlsx, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            if n == target:
                continue
            d = data_map[n]
            if n.endswith(".rels"):
                d = d.replace(old_base.encode(), new_base.encode())
            if n == "[Content_Types].xml" and b'Extension="png"' not in d:
                d = d.replace(b"</Types>", b'<Default Extension="png" ContentType="image/png"/></Types>')
            zout.writestr(n, d)
        zout.writestr("xl/media/" + new_base, new_png)
    return 1


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    root = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
    c = build(root)
    print("通用空白模板:", DST)
    print(f"  清 FAI 动态格: {c['FAI']}")
    print(f"  清 OLE 表标签: {c['标签']}")
    print(f"  清样品照片残照: {c['照片']} 张")
