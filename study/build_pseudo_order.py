# -*- coding: utf-8 -*-
"""M2.0 伪真单构造器:从成品 golden 拆出"操作员手上的原始输入",破循环验证。

产出 本单输入/pseudo/<code>/:
  drawing/<code>_drawing.pdf   —— 生图纸(供 qwen-vl 抽 FAI/品号)
  materials/*.pdf              —— 材料 PDF 堆: 原始文件名、扁平、**剥掉表归属**、按文件名去重
                                  (操作员手上每份只有一个拷贝; 扇出=分类器的活, 不在此泄露)
  photos/*.png                 —— 样品照片
  _groundtruth.json            —— 仅供打分: {file_slots(真表归属), fai(真尺寸), golden路径}
                                  **管线严禁读取**, 只有 scorer 用来判分
"""
import os, sys, re, json, zipfile, shutil

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
import openpyxl
from hitl.ole_assemble import extract_embedded_pdf, original_filename
from study.case_data import _distinct_bins_per_sheet, extract_photos

CASES_DIR = r"案例材料\承认书\参考用承诺书集"
OUTROOT = r"本单输入\pseudo"


def slot(sheet):
    s = sheet.strip()
    if "材质成分" in s or "展开" in s: return "材质表"
    if "材质证明" in s: return "材质证明"
    if "UL" in s.upper(): return "UL"
    if "部件" in s: return "部件承认"
    if "信赖" in s or "信賴" in s: return "信赖性"
    if "包装" in s or "包裝" in s: return "包装"
    if "出货" in s or "出貨" in s or "检验" in s: return "出货"
    if "图纸" in s: return "图纸"
    return s[:6]


def drawing_bin(g):
    z = zipfile.ZipFile(g); names = set(z.namelist())
    sheets = re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"',
                        z.read("xl/workbook.xml").decode("utf-8", "ignore"))
    rel = dict(re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"',
               z.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")))
    for nm, rid in sheets:
        if "图纸" not in nm: continue
        wsf = rel[rid]; wsf = ("xl/" + wsf) if not wsf.startswith("xl/") else wsf
        relf = f"xl/worksheets/_rels/{os.path.basename(wsf)}.rels"
        if relf in names:
            bins = [os.path.basename(t) for t in re.findall(r'Target="([^"]+)"',
                    z.read(relf).decode("utf-8", "ignore")) if "embeddings/oleObject" in t]
            if bins: return bins[0]
    return None


def golden_fai(g):
    wb = openpyxl.load_workbook(g, data_only=True)
    fws = [wb[s] for s in wb.sheetnames if "FAI" in s]
    if not fws: return []
    f = fws[0]; out = []
    for r in range(9, 45):
        try:
            b, c, d = float(f.cell(r, 2).value), float(f.cell(r, 3).value), float(f.cell(r, 4).value)
        except (TypeError, ValueError):
            continue
        out.append([c, round(d - c, 3), round(c - b, 3)])
    return out


def _safe(nm):
    return re.sub(r'[\\/:*?"<>|]', "_", nm).strip()


def build_pseudo(golden, outroot):
    code = re.sub(r"[^A-Za-z0-9]", "", os.path.basename(golden))[:10]
    out = os.path.join(outroot, code)
    if os.path.exists(out):
        shutil.rmtree(out)
    os.makedirs(os.path.join(out, "drawing"), exist_ok=True)
    os.makedirs(os.path.join(out, "materials"), exist_ok=True)

    # ① 生图纸
    db = drawing_bin(golden)
    draw_ok = False
    if db:
        try:
            extract_embedded_pdf(golden, db, os.path.join(out, "drawing", f"{code}_drawing.pdf"))
            draw_ok = True
        except Exception:
            pass

    # ② 材料堆: 收集 文件名→真表归属(可多), 按名去重
    sb = _distinct_bins_per_sheet(golden)
    file_slots, file_bin = {}, {}
    for sheet, bins in sb.items():
        sl = slot(sheet)
        if sl == "图纸":
            continue
        for b in bins:
            nm = _safe(original_filename(golden, b) or b.replace(".bin", ".pdf"))
            if not nm.lower().endswith(".pdf"):
                nm += ".pdf"
            file_slots.setdefault(nm, set()).add(sl)
            file_bin.setdefault(nm, b)   # 任取一个 bin 抽(扇出的多份源相同)
    extracted = 0
    for nm, b in file_bin.items():
        try:
            extract_embedded_pdf(golden, b, os.path.join(out, "materials", nm))
            extracted += 1
        except Exception:
            file_slots.pop(nm, None)

    # ③ 照片
    photos = extract_photos(golden, os.path.join(out, "photos"))

    # ④ ground truth(仅打分用, 管线禁读)
    gt = {"golden": os.path.abspath(golden), "code": code,
          "file_slots": {k: sorted(v) for k, v in file_slots.items()},
          "fai": golden_fai(golden)}
    with open(os.path.join(out, "_groundtruth.json"), "w", encoding="utf-8") as f:
        json.dump(gt, f, ensure_ascii=False, indent=1)

    fanout = sum(1 for v in file_slots.values() if len(v) > 1)
    return {"code": code, "draw": draw_ok, "materials": extracted,
            "photos": len(photos), "fanout": fanout,
            "slots": {s: sum(1 for v in file_slots.values() if s in v)
                      for s in ["材质表", "材质证明", "部件承认", "UL", "信赖性", "包装", "出货"]}}


def main():
    import io, glob
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    cases = sorted([p for p in glob.glob(os.path.join(CASES_DIR, "*.xlsx"))
                    if not os.path.basename(p).startswith("~$")])
    print(f"=== M2.0 伪真单构造 | {len(cases)}案 ===\n")
    for g in cases:
        r = build_pseudo(g, OUTROOT)
        sl = " ".join(f"{k}{v}" for k, v in r["slots"].items() if v)
        print(f"{'✅' if r['draw'] else '⚠️'} {r['code']}: 图纸{'✓' if r['draw'] else '✗'} "
              f"材料{r['materials']}份(扇出{r['fanout']}) 照片{r['photos']} | 真值[{sl}]")
    print(f"\n产出 本单输入/pseudo/<code>/ (drawing + materials扁平 + photos + _groundtruth.json)")


if __name__ == "__main__":
    main()
