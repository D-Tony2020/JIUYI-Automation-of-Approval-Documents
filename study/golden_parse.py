# -*- coding: utf-8 -*-
"""把一份承认书的『7.材质成分展开表』解析成结构化 golden（按材质分组）。

供 BOM 抽取盲测对标。纯本地、不外发。
材质表列：B零件 C供应商 D材质类别 E材质 G成份名 H CAS J重量% M-V RoHS十项 W日期 X报告号。
材质 = 列E 的合并段；成份 = 列G 非空的行；RoHS/报告 = 该报告块首行（合并锚点）。
"""
import openpyxl

ROHS_KEYS = ["Pb", "Cd", "Hg", "Cr6+", "PBBs", "PBDEs", "DEHP", "DBP", "BBP", "DIBP"]


def find_mat_sheet(wb):
    for s in wb.sheetnames:
        if "材质成分" in s:
            return wb[s]
    raise ValueError("未找到材质成分展开表")


def _merge_of(merges, row, col):
    """返回包含(row,col)的合并区(min_row,max_row,min_col)，无则 None。"""
    for mr in merges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr
    return None


def _anchor_val(ws, merges, row, col):
    """合并区取左上角值；非合并取本格。"""
    mr = _merge_of(merges, row, col)
    if mr:
        return ws.cell(mr.min_row, mr.min_col).value
    return ws.cell(row, col).value


def parse_golden(xlsx_path):
    """→ [{零件,材质类别,材质,_mat_first,components:[{成份名,CAS,重量%,_row}],
          blocks:[{_first,rohs:{},报告号,报告日期}]}]，按材质(列E合并段)分组。"""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = find_mat_sheet(wb)
    merges = list(ws.merged_cells.ranges)
    mats = {}
    order = []
    for r in range(14, ws.max_row + 1):
        name = ws.cell(r, 7).value  # G 成份名
        if name is None or str(name).strip() == "":
            continue
        mr5 = _merge_of(merges, r, 5)
        mat_first = mr5.min_row if mr5 else r          # 材质段首行=分组键
        mat_name = _anchor_val(ws, merges, r, 5)
        if mat_first not in mats:
            mats[mat_first] = {
                "零件": _anchor_val(ws, merges, r, 2),
                "供应商": _anchor_val(ws, merges, r, 3),
                "材质类别": _anchor_val(ws, merges, r, 4),
                "材质": mat_name,
                "_mat_first": mat_first,
                "components": [],
                "blocks": {},
            }
            order.append(mat_first)
        m = mats[mat_first]
        m["components"].append({
            "成份名": str(name).strip(),
            "CAS": str(ws.cell(r, 8).value or "").strip(),
            "重量%": str(ws.cell(r, 10).value if ws.cell(r, 10).value is not None else "").strip(),
            "_row": r,
        })
        # 报告块（列X 报告号合并段）→ RoHS/报告
        mrX = _merge_of(merges, r, 24)
        blk_first = mrX.min_row if mrX else r
        if blk_first not in m["blocks"]:
            m["blocks"][blk_first] = {
                "_first": blk_first,
                "rohs": {ROHS_KEYS[i]: str(_anchor_val(ws, merges, r, 13 + i) or "").strip()
                         for i in range(10)},
                "报告号": str(_anchor_val(ws, merges, r, 24) or "").strip(),
                "报告日期": str(_anchor_val(ws, merges, r, 23) or "").strip(),
            }
    out = []
    for k in order:
        m = mats[k]
        m["blocks"] = list(m["blocks"].values())
        out.append(m)
    return out


if __name__ == "__main__":
    import sys, io, os, json
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    p = sys.argv[1] if len(sys.argv) > 1 else \
        r"案例材料\承认书\承认书\做好的承认书\YY60039403 (J00016372) 承认书.xlsx"
    mats = parse_golden(p)
    print(f"解析 {os.path.basename(p)[:14]}: {len(mats)} 个材质, "
          f"{sum(len(m['components']) for m in mats)} 个成份")
    for m in mats:
        print(f"  [{m['材质']}] 零件={m['零件']} 成份×{len(m['components'])} 报告块×{len(m['blocks'])} "
              f"行{m['_mat_first']}  CAS={[c['CAS'] for c in m['components']][:6]}")
