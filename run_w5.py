# -*- coding: utf-8 -*-
"""W5 FAI 执行线：治病清旧值 + 表头 + 图纸尺寸→规格 + 实测红槽（公式保留）。

build_upto(4)=封面+送样+材质表+FAI；空白参照=build_upto(3)(无FAI)。
自检：规格 B/C/D==图纸尺寸换算(对golden)、E-H公式保留、表头、实测红槽、旧值已清。
"""
import os
import sys
import io
import json
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.build import build_upto
from hitl.harness import assert_no_external_links
from hitl import fai

ROOT = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(ROOT, "模板", "承认书空白模板_治病.xlsx")
GOLDEN = os.path.join(ROOT, "案例材料", "承认书", "承认书", "做好的承认书",
                      "YY60039403 (J00016372) 承认书.xlsx")
BOM_JSON = os.path.join(ROOT, "hitl", "data", "demo_bom_YY60039403.json")
OUTDIR = os.path.join(ROOT, "产出留档", "W5-FAI")
BLANK = os.path.join(OUTDIR, "FAI_空白参照__00.xlsx")
OUT = os.path.join(OUTDIR, "FAI_规格+红槽__01.xlsx")
SHEET = fai.FAI_SHEET

# 本单 YY60039403 图纸 4 尺寸（HITL 录入）
DIMENSIONS = [(98, 5), (60, 5), (28, 3), (2, 0.5)]
DATA = {
    "drawing_meta": {"名称": "SB120420BLCNR009导线", "品号": "YY60039403", "版本": "A01"},
    "product": {"材料名称": "导线", "填表日期": datetime.date(2026, 6, 24)},
    "bom": json.load(open(BOM_JSON, encoding="utf-8")),
    "dimensions": DIMENSIONS,
}


def main():
    build_upto(TPL, BLANK, DATA, upto=3)  # 不含 FAI
    build_upto(TPL, OUT, DATA, upto=4)    # 含 FAI

    out_ws = openpyxl.load_workbook(OUT)[SHEET]
    errs = fai.selfcheck_fai(out_ws, DIMENSIONS)
    assert_no_external_links(OUT)

    # 对 golden 核规格 B9:D12
    gnames = openpyxl.load_workbook(GOLDEN).sheetnames
    gold = openpyxl.load_workbook(GOLDEN, data_only=True)[
        [s for s in gnames if s.strip() == SHEET.strip()][0]]
    for r in range(9, 9 + len(DIMENSIONS)):
        for c in (2, 3, 4):
            gv, ov = gold.cell(r, c).value, out_ws.cell(r, c).value
            if gv != ov:
                errs.append(f"规格 r{r}c{c} 对golden不一致: golden={gv} 我={ov}")
    # 表头
    if out_ws["C2"].value != "导线":
        errs.append(f"C2 品类={out_ws['C2'].value!r}")
    if out_ws["U2"].value != "2026.06.24":
        errs.append(f"U2 日期={out_ws['U2'].value!r}")
    # 旧值已清（item 5-30 的规格空 = 模板里曾有？本单仅4项，B13应空）
    if out_ws["B13"].value not in (None, ""):
        errs.append(f"item5 规格未清: B13={out_ws['B13'].value!r}")
    # 实测红槽（I9 红）
    f9 = out_ws["I9"].fill
    red = f9.fgColor.rgb if f9 and f9.patternType == "solid" else None
    if red != "00FF0000":
        errs.append(f"I9 实测未标红槽: {red!r}")

    status = "PASS" if not errs else "FAIL"
    os.makedirs(OUTDIR, exist_ok=True)
    with open(os.path.join(OUTDIR, "_manifest.txt"), "a", encoding="utf-8") as f:
        f.write(f"FAI_规格+红槽__01\t规格4行 C2={out_ws['C2'].value} U2={out_ws['U2'].value}\t{status}\n")

    print("空白参照:", BLANK)
    print("填入成品:", OUT)
    print("表头: C2(品类)=%s  U2(日期)=%s  I2(图号公式)=%s  O2(版本公式)=%s" % (
        out_ws["C2"].value, out_ws["U2"].value, out_ws["I2"].value, out_ws["O2"].value))
    print("规格(图纸尺寸驱动):")
    for r in range(9, 9 + len(DIMENSIONS)):
        print(f"   item{r-8}: B={out_ws.cell(r,2).value} C={out_ws.cell(r,3).value} D={out_ws.cell(r,4).value}  (E公式={out_ws.cell(r,5).value})")
    if errs:
        print("❌ 自测失败：")
        for e in errs:
            print("   -", e)
        sys.exit(1)
    print("✅ 自测通过：规格==图纸(对golden) + CP/CPK公式保留 + 表头 + 实测红槽 + 旧值已清 + 无外链")


if __name__ == "__main__":
    main()
