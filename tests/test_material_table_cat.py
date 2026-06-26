# -*- coding: utf-8 -*-
"""#4A 材质类别子分组: inject_data 把 D列(材质类别)按"零件内连续同类别"组合并,
E/F(材质/重量)按材质。修 to_inject_bom/stage2_to_nested_bom 坍缩 bug(多类别零件被坍成只首类别)。"""
import os
import sys

import openpyxl

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
sys.path.insert(0, ROOT)

from hitl.material_table import inject_data, DATA_TOP


def _blk(name, cas, w="0.5"):
    return {"成份": [{"成份名称": name, "CAS": cas, "重量%": w}], "RoHS": {}, "报告编号": "", "报告日期": ""}


BOM = [
    {"零件": "导线", "供应商": "正崴", "materials": [
        {"材质类别": "线材", "材质": "PVC", "blocks": [_blk("a", "1-1-1")]},
        {"材质类别": "线材", "材质": "镀锡铜", "blocks": [_blk("b", "2-2-2")]},
    ]},
    {"零件": "胶座端子", "供应商": "联和", "materials": [   # 一零件两类别(golden 实况)
        {"材质类别": "胶座", "材质": "PA66", "blocks": [_blk("c", "3-3-3")]},
        {"材质类别": "端子", "材质": "磷青铜", "blocks": [_blk("d", "4-4-4")]},
    ]},
]


def _run():
    wb = openpyxl.Workbook()
    ws = wb.active
    inject_data(ws, BOM, DATA_TOP)
    return ws, {str(m) for m in ws.merged_cells.ranges}


def test_同类别多材质_D合并一格():
    ws, merges = _run()
    # 导线: PVC(14)+镀锡铜(15) 同类别"线材" → D14:D15 合并成一格
    assert ws.cell(14, 4).value == "线材"
    assert "D14:D15" in merges


def test_材质E按材质不合并():
    ws, _ = _run()
    assert ws.cell(14, 5).value == "PVC" and ws.cell(15, 5).value == "镀锡铜"   # E 各材质各自


def test_一零件多类别_D不跨合并():
    ws, merges = _run()
    # 胶座端子: 胶座(16)/端子(17) 不同类别 → D 不跨合并, 各显各
    assert ws.cell(16, 4).value == "胶座" and ws.cell(17, 4).value == "端子"
    assert "D16:D17" not in merges


def test_零件供应商零件级合并():
    ws, merges = _run()
    assert ws.cell(14, 2).value == "导线" and ws.cell(14, 3).value == "正崴"
    assert "B14:B15" in merges and "C14:C15" in merges     # 零件/供应商跨零件合并
