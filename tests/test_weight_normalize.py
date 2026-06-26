# -*- coding: utf-8 -*-
"""重量% 归一(范围取中+÷100)与材质表写入(float+'0.00%'格式)。

修"多除100": ①范围值取中(非低端) ②写单元格用float+百分比格式(否则字符串裸显小数像被多除)。
两份 normalize_weight(spike/assemble + hitl/material_table)须统一。
"""
import os
import sys

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..")
sys.path.insert(0, ROOT)
sys.path.insert(0, os.path.join(ROOT, "spike"))      # spike/assemble 内用 'from schemas import'

import openpyxl

from spike.assemble import normalize_weight as nw_spike
from hitl.material_table import normalize_weight as nw_mat, _put_weight


def test_范围取中对齐golden():
    # golden 磷青铜: 锌 0.00035, 镍 0.00058(范围中值÷100)
    assert nw_spike("0.03-0.04%") == "0.00035"          # 中值 0.035% → 0.00035(非低端0.0003)
    assert nw_spike("0.054-0.063%") == "0.000585"
    assert nw_spike("6.2 %") == "0.062"
    assert nw_spike("56.4%") == "0.564"


def test_范围内减号不当负号():
    v = float(nw_spike("0.0026-0.005%"))                # 中值(0.0026+0.005)/2÷100≈3.8e-05
    assert 3.5e-05 < v < 4.0e-05                         # 正数, 不是 (0.0026-0.005) 的负数


def test_余量与上限原样():
    assert nw_spike("余量") == "余量" and nw_spike("balance") == "余量"
    assert nw_spike("＜0.005%") == "＜0.005%" and nw_spike("<3") == "<3"


def test_两份normalize统一():
    for raw in ("0.03-0.04%", "6.2 %", "56.4%", "余量", "＜0.005%"):
        assert nw_spike(raw) == nw_mat(raw)


def test_put_weight数值带百分比格式():
    wb = openpyxl.Workbook(); ws = wb.active
    _put_weight(ws, 1, "0.062")
    assert ws.cell(1, 10).value == 0.062 and ws.cell(1, 10).number_format == "0.00%"  # float→渲染6.20%


def test_put_weight余量保字符串():
    wb = openpyxl.Workbook(); ws = wb.active
    _put_weight(ws, 1, "余量")
    assert ws.cell(1, 10).value == "余量"                # 非数值原样, 不强转
