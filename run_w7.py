# -*- coding: utf-8 -*-
"""W7 样品照片执行线：照片数驱动布局。给定 N 张(2-4)上传照 → 摆进 Excel。

照片来源=前端UI上传(微信→剪贴板→粘贴)；本脚本用 golden 抽出的成品照当 demo 源，
产出 N=2/3/4 三种布局验证。自检：清残照留LOGO、放入N张、不重叠、无外链。
"""
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.harness import assert_no_external_links
from hitl import sample_photo

ROOT = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(ROOT, "模板", "承认书空白模板_治病.xlsx")
GOLDEN = os.path.join(ROOT, "案例材料", "承认书", "承认书", "做好的承认书",
                      "YY60039403 (J00016372) 承认书.xlsx")
PHOTO_DIR = os.path.join(ROOT, "本单输入", "photos_demo")
OUTDIR = os.path.join(ROOT, "产出留档", "W7-样品照片")
SHEET = sample_photo.SHEET


def extract_demo_photos():
    """从 golden 抽出成品照(非LOGO)当 demo 上传源。"""
    os.makedirs(PHOTO_DIR, exist_ok=True)
    gws = openpyxl.load_workbook(GOLDEN)[SHEET]
    paths = []
    for i, im in enumerate(gws._images):
        if sample_photo._anchor_col(im) == 0:
            continue
        p = os.path.join(PHOTO_DIR, f"demo_{i}.png")
        with open(p, "wb") as f:
            f.write(im._data())
        paths.append(p)
    return paths


def main():
    os.makedirs(OUTDIR, exist_ok=True)
    photos = extract_demo_photos()
    print(f"demo 成品照源: {len(photos)} 张（从 golden 抽）")

    all_ok = True
    for n in (2, 3, 4):
        src = (photos * 3)[:n]                    # 不足则复用，凑 N
        wb = openpyxl.load_workbook(TPL)
        ws = wb[SHEET]
        placed = sample_photo.fill_sample_photo(ws, src)
        out = os.path.join(OUTDIR, f"样品照片_{n}张__01.xlsx")
        if os.path.exists(out):
            os.remove(out)
        wb.save(out)

        rews = openpyxl.load_workbook(out)[SHEET]
        errs = sample_photo.selfcheck_sample_photo(rews, n)
        assert_no_external_links(out)
        reopen_photos = sum(1 for im in rews._images if sample_photo._anchor_col(im) != 0)
        lay = sample_photo.photo_layout(n)
        shape = f"{n//2}行×2" + ("+1居中" if n % 2 else "")
        st = "PASS" if not errs else "FAIL: " + "; ".join(errs)
        print(f"  N={n}: 放入{placed}张 复开见{reopen_photos}张 布局[{shape}] "
              f"位置={[(round(x,1),r) for _, x, r, _ in lay]}  {st}")
        all_ok = all_ok and not errs

    with open(os.path.join(OUTDIR, "_manifest.txt"), "a", encoding="utf-8") as f:
        f.write(f"样品照片 N=2/3/4 布局\t{'PASS' if all_ok else 'FAIL'}\n")
    if not all_ok:
        sys.exit(1)
    print("✅ 自测通过：N=2/3/4 布局自适应 + 留LOGO + 不重叠 + 无外链")


if __name__ == "__main__":
    main()
