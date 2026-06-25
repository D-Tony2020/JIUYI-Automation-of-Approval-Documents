# -*- coding: utf-8 -*-
"""W8 校验闸：齐套(漏嵌) / 有效期预警 / 未填禁导出 / 溯源面板。纯本地。

- 齐套：各 OLE 表实测槽位数 vs BOM 应有数(计数规律)；实测<应有=漏嵌。
- 有效期：材质表报告日期 + 有效期 vs 今日；过期/即将过期=预警(不硬拦，老板定)。
- 未填：样品照片<2 等必填空槽=拦截(禁导出)。
- 溯源：每材质→报告编号/日期/供应商，导出前可核。
"""
import datetime
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from study.ole_structure import ole_count_per_sheet
from study.embed_structure import embed_group_count
from study.golden_parse import parse_golden
from hitl.sample_photo import SHEET as PHOTO_SHEET, _anchor_col

import openpyxl


def _parse_date(s):
    s = str(s or "").strip()
    for sep in (".", "-", "/"):
        parts = s.split(sep)
        if len(parts) == 3 and parts[0].isdigit():
            try:
                return datetime.date(int(parts[0]), int(parts[1]), int(parts[2]))
            except ValueError:
                pass
    return None


def check_completeness(xlsx, materials):
    """各 OLE 表 实测槽位 vs 应有数。返回漏嵌 [(表, 实测, 应有)]。"""
    exp = embed_group_count(materials)
    act = ole_count_per_sheet(xlsx)
    issues = []
    for sheet in ("材质证明书", "部件承认书", "UL证明", "信赖性"):
        a, e = act.get(sheet, 0), exp.get(sheet, 0)
        if a < e:
            issues.append((sheet, a, e))
    return issues


def check_validity(xlsx, today, valid_months=12, warn_days=90):
    """材质表报告日期 + 有效期 vs 今日。返回 [(材质, 日期, 状态)]，状态∈过期/即将过期。"""
    materials = parse_golden(xlsx)
    out = []
    for m in materials:
        for b in m["blocks"]:
            d = _parse_date(b.get("报告日期"))
            if not d:
                continue
            exp = datetime.date(d.year + valid_months // 12, d.month, d.day)
            if exp < today:
                out.append((m["材质"].strip(), b.get("报告日期"), "过期"))
            elif (exp - today).days <= warn_days:
                out.append((m["材质"].strip(), b.get("报告日期"), f"{(exp-today).days}天内到期"))
    return out


def check_required(xlsx, materials):
    """必填空槽。返回拦截项 list。"""
    miss = []
    wb = openpyxl.load_workbook(xlsx)
    ws = [wb[s] for s in wb.sheetnames if s.strip() == PHOTO_SHEET.strip()]
    nphoto = sum(1 for im in ws[0]._images if _anchor_col(im) != 0) if ws else 0
    if nphoto < 2:
        miss.append(f"样品照片仅 {nphoto} 张（需≥2，待 UI 上传）")
    return miss


def trace_panel(xlsx):
    """溯源：每材质→报告编号/日期/供应商。"""
    rows = []
    for m in parse_golden(xlsx):
        for b in m["blocks"]:
            rows.append({"零件": m["零件"], "材质": m["材质"].strip(),
                         "报告编号": b.get("报告号", ""), "报告日期": b.get("报告日期", "")})
    return rows


def gate(xlsx, materials, today):
    """汇总校验。返回 (can_export, report)。未填→禁导出；有效期/齐套→预警。"""
    comp = check_completeness(xlsx, materials)
    valid = check_validity(xlsx, today)
    req = check_required(xlsx, materials)
    can_export = not req            # 未填硬拦；齐套/有效期仅预警
    return can_export, {"漏嵌": comp, "有效期预警": valid, "未填拦截": req}
