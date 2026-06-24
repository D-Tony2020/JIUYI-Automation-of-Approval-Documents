# -*- coding: utf-8 -*-
"""W0 执行台：模板加载 / 合并区感知写入 / 落盘 / 读回断言。逐表 HITL 复用。

设计原则：
- 只 save 一次（防 openpyxl 往返复制 media 膨胀）。
- 封面这类无 OLE 的表用 openpyxl 即可；含 OLE 的装配表（W6）走 COM，不在此。
- 写入合并单元格自动落到左上锚点格。
"""
import os
import hashlib
import zipfile
import openpyxl


def load_template(template_path):
    """加载黄金模板，保留公式与图片（data_only=False）。"""
    return openpyxl.load_workbook(template_path)


def _merged_anchor(ws, coord):
    """若 coord 落在某合并区内，返回该区左上格坐标；否则原样返回。"""
    cell = ws[coord]
    for mr in ws.merged_cells.ranges:
        min_c, min_r, max_c, max_r = mr.bounds
        if min_r <= cell.row <= max_r and min_c <= cell.column <= max_c:
            return ws.cell(min_r, min_c).coordinate
    return coord


def write_cell(ws, coord, value):
    """写入单元格；合并区自动写左上锚点格。"""
    ws[_merged_anchor(ws, coord)] = value


def read_cell(ws, coord):
    """读取单元格；合并区从左上锚点格读。"""
    return ws[_merged_anchor(ws, coord)].value


def save_output(wb, out_path, dedupe=True):
    """落盘（只 save 一次）+ 合并重复图片。自动建目录。"""
    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    wb.save(out_path)
    if dedupe:
        dedupe_media(out_path)
    return out_path


def dedupe_media(xlsx_path):
    """合并 openpyxl 保存时按引用重复写入的相同图片（按内容哈希），重指 rels。

    根因：多 sheet 共享同一 Logo，openpyxl 各存一份导致 media 膨胀（6→18）。
    返回 (前, 后) media 数。
    """
    with zipfile.ZipFile(xlsx_path) as z:
        parts = {n: z.read(n) for n in z.namelist()}
    media = {n: d for n, d in parts.items() if n.startswith("xl/media/")}
    before = len(media)

    # 内容哈希 → 规范名（确定性取排序最小）
    by_hash = {}
    for n in sorted(media):
        by_hash.setdefault(hashlib.md5(media[n]).hexdigest(), n)
    dup_map = {}  # 重复 basename → 规范 basename
    for n in media:
        canon = by_hash[hashlib.md5(media[n]).hexdigest()]
        if canon != n:
            dup_map[os.path.basename(n)] = os.path.basename(canon)
    if not dup_map:
        return before, before

    # 重写所有 .rels 对重复图片的引用（带 .png/.jpeg 后缀，边界安全）
    for n in list(parts):
        if n.endswith(".rels"):
            s = parts[n].decode("utf-8")
            for dup, canon in dup_map.items():
                if dup in s:
                    s = s.replace(dup, canon)
            parts[n] = s.encode("utf-8")
    # 删除重复 media
    for n in list(parts):
        if n.startswith("xl/media/") and os.path.basename(n) in dup_map:
            del parts[n]

    with zipfile.ZipFile(xlsx_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for n, data in parts.items():
            zout.writestr(n, data)
    after = sum(1 for n in parts if n.startswith("xl/media/"))
    return before, after


def emit_blank_reference(template_path, out_path):
    """产出该表《空白参照》：模板过同一落盘管线但不填任何值。

    手测前与填入成品做 diff 对照：确认空白初始化正确 + 填入定位精确。
    """
    wb = load_template(template_path)
    return save_output(wb, out_path)


def diff_cells(blank_path, filled_path, sheet_name):
    """单元格级 diff：返回 [(坐标, 空白值, 填入值)]。证明只动了该动的格。"""
    sb = openpyxl.load_workbook(blank_path)[sheet_name]
    sf = openpyxl.load_workbook(filled_path)[sheet_name]
    maxr = max(sb.max_row, sf.max_row)
    maxc = max(sb.max_column, sf.max_column)
    out = []
    for r in range(1, maxr + 1):
        for c in range(1, maxc + 1):
            vb = sb.cell(r, c).value
            vf = sf.cell(r, c).value
            if vb != vf:
                out.append((sf.cell(r, c).coordinate, vb, vf))
    return out


def assert_no_external_links(xlsx_path):
    """zip 级断言：产物不含外链（否则 Excel 打开报错）。"""
    with zipfile.ZipFile(xlsx_path) as z:
        bad = [n for n in z.namelist() if n.startswith("xl/externalLinks")]
    if bad:
        raise AssertionError(f"产物含外链 {bad}")


def count_media(xlsx_path):
    """zip 级：media 图片数（防膨胀回归用）。"""
    with zipfile.ZipFile(xlsx_path) as z:
        return sum(1 for n in z.namelist() if n.startswith("xl/media/"))
