# -*- coding: utf-8 -*-
"""多份承认书取交集 + 空白模板状态 → 判定每格 静态/动态（Phase A 增强工具）。

方法（老板提出）：跨 N 份不同产品承认书，恒定=静态候选、变动=动态。
再结合空白模板"该格是否预填"做组合判据：
  varies              → 动态(必填)
  constant + 模板空   → 动态·定值（如 D12=导线，仍留 fill 以备扩品类）
  constant + 模板已填 → 静态
  varies + 模板已填   → ⚠️模板初始化错（把会变的值写死了）—— 必须修模板

用法：每张表 Phase A 调 classify(sheet, blank_template) 得到该表 动态格集合，
作为"填哪些格"的客观依据，而非凭单份样本或肉眼推断。
"""
import glob
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter


def _norm(v):
    """折叠空白用于"是否变动"比较，去掉纯格式（空格/缩进）噪声。"""
    return re.sub(r"\s+", "", str(v)) if v not in (None, "") else ""

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SAMPLES_DIR = os.path.join(_ROOT, "案例材料", "承认书", "参考用承诺书集")
BLANK_TEMPLATE = os.path.join(_ROOT, "模板", "承认书空白模板_治病.xlsx")


def _sample_paths():
    return sorted(p for p in glob.glob(os.path.join(SAMPLES_DIR, "*.xlsx"))
                  if not os.path.basename(p).startswith("~$"))


def _match(wb, sheet):
    return next((s for s in wb.sheetnames if s.strip() == sheet.strip()), None)


def _grid(ws, max_row, max_col):
    return {(r, c): ws.cell(r, c).value
            for r in range(1, max_row + 1) for c in range(1, max_col + 1)}


def classify(sheet, blank_template=BLANK_TEMPLATE, max_row=80, max_col=40):
    """返回 (结果dict, 样本数)。结果: {coord: {verdict, 样本值[], 模板值}}。"""
    grids = []
    for p in _sample_paths():
        wb = openpyxl.load_workbook(p, data_only=True)
        nm = _match(wb, sheet)
        if nm:
            grids.append(_grid(wb[nm], max_row, max_col))

    blank = {}
    if blank_template:
        wb = openpyxl.load_workbook(blank_template)
        nm = _match(wb, sheet)
        if nm:
            blank = _grid(wb[nm], max_row, max_col)

    coords = set().union(*[set(g) for g in grids]) if grids else set()
    out = {}
    for (r, c) in sorted(coords):
        vals = [g.get((r, c)) for g in grids]
        if all(v in (None, "") for v in vals):
            continue
        coord = f"{get_column_letter(c)}{r}"
        varies = len({_norm(v) for v in vals}) > 1   # 折叠空白后比较，去格式噪声
        tmpl = blank.get((r, c))
        tmpl_filled = tmpl not in (None, "")
        tmpl_formula = isinstance(tmpl, str) and tmpl.startswith("=")
        if tmpl_formula:
            verdict = "公式·自动"          # 模板公式自动派生(如 =封面!D14)，非错
        elif varies and tmpl_filled:
            verdict = "模板初始化错"        # 模板写死了会变的值 → 须修模板
        elif varies:
            verdict = "动态"
        elif tmpl_filled:
            verdict = "静态"
        else:
            verdict = "动态·定值"
        out[coord] = {"verdict": verdict, "vals": vals, "tmpl": tmpl}
    return out, len(grids)


def dynamic_coords(sheet, **kw):
    """该表需工具填的格集合（动态 + 动态·定值），= '填哪些格'的客观依据。"""
    res, _ = classify(sheet, **kw)
    return {c for c, d in res.items() if d["verdict"] in ("动态", "动态·定值")}


if __name__ == "__main__":
    import sys
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    sheets = sys.argv[1:] or ["封面", "1.送样承认书目录"]
    for sh in sheets:
        res, n = classify(sh)
        print(f"\n===== 「{sh}」 样本{n} =====")
        for verdict in ("模板初始化错", "公式·自动", "动态", "动态·定值"):
            hits = {c: d for c, d in res.items() if d["verdict"] == verdict}
            if hits:
                print(f"  [{verdict}] {len(hits)}:")
                for c, d in hits.items():
                    print(f"    {c}: 样本={[str(v)[:14] for v in d['vals']]} 模板={d['tmpl']!r}")
        nst = sum(1 for d in res.values() if d["verdict"] == "静态")
        print(f"  [静态] {nst} 个（略）")
