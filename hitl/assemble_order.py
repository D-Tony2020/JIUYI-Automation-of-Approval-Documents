# -*- coding: utf-8 -*-
"""段二装配(M2.4 dry-run / M2.5 真导出): stage2_bom 文件↔材质链 → 终态承认书。

流程: 段一 build_upto 填格(材质表按 nested_bom 变行重构) → 段二 embed_many COM 嵌 OLE → verify_open。
最高不变式: 段一 build_upto 的 bom 与 build_specs 的 nested 必须**同源同序**(都来自
stage2_to_nested_bom 的同一调用), 否则材质表 OLE 落错行。
plan_only 不碰 COM(可单测/dry-run); assemble 走 COM(WPS), 慢且依赖装机。
"""
import datetime
import os

import openpyxl

from hitl.build import build_upto
from hitl.ole_assemble import count_ole, embed_many, make_icon, verify_open
from hitl.placement_plan import build_specs, stage2_to_nested_bom
from hitl import sample_photo

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BLANK = os.path.join(ROOT, "模板", "承认书空白模板_通用.xlsx")


def plan_only(stage2_bom, drawing_meta, dimensions, materials_dir, drawing_pdf, outdir, blank=BLANK):
    """段一 cell + build_specs(无 COM/无 icon)。返回 (cell_path, specs, nested)。

    段一与放置同源: nested 来自 stage2_to_nested_bom, build_upto 用它重构材质表行结构。
    """
    os.makedirs(outdir, exist_ok=True)
    nested, _ordered = stage2_to_nested_bom(stage2_bom.get("materials", []))
    name = drawing_meta.get("名称") or drawing_meta.get("品类") or "导线"
    data = {"drawing_meta": {"名称": name, "品号": drawing_meta.get("品号", ""),
                             "版本": drawing_meta.get("版本", "")},
            "product": {"材料名称": name, "填表日期": datetime.date(2026, 6, 25)},
            "bom": nested, "dimensions": dimensions or []}
    cell = os.path.join(outdir, "_cell.xlsx")
    build_upto(blank, cell, data, upto=4, highlight=False)
    sheet_names = openpyxl.load_workbook(cell, read_only=True).sheetnames
    specs, _nested2, _ = build_specs(dict(stage2_bom), sheet_names, materials_dir, drawing_pdf)
    return cell, specs, nested


def count_specs_by_sheet(specs):
    """specs → {表名: OLE 数}(spec 级核对, 无 COM)。"""
    out = {}
    for s in specs:
        out[s["sheet"]] = out.get(s["sheet"], 0) + 1
    return out


def assemble(stage2_bom, drawing_meta, dimensions, materials_dir, drawing_pdf, out_xlsx,
             outdir, photos=None, blank=BLANK):
    """全装配(走 COM): 段一+照片 → make_icon → embed_many → verify_open。返回 dict。

    失败不删 cell(段一成果留存)。调用方应在子进程+超时下跑(WPS 可能卡)。
    """
    cell, specs, nested = plan_only(stage2_bom, drawing_meta, dimensions,
                                    materials_dir, drawing_pdf, outdir, blank)
    if photos:
        wb = openpyxl.load_workbook(cell)
        try:
            sample_photo.fill_sample_photo(wb[sample_photo.SHEET], photos)
        except Exception:
            pass
        wb.save(cell)
    icondir = os.path.join(outdir, "icons")
    os.makedirs(icondir, exist_ok=True)
    for s in specs:
        s["icon"] = make_icon(s["pdf"], os.path.join(icondir, os.path.basename(s["pdf"]) + ".png"))
    if os.path.exists(out_xlsx):
        os.remove(out_xlsx)
    embed_many(cell, out_xlsx, specs)
    try:
        opens = verify_open(out_xlsx).get("_total", -1)
    except Exception:
        opens = -1
    return {"out": out_xlsx, "specs": len(specs), "ole": count_ole(out_xlsx), "opens": opens,
            "by_sheet": count_specs_by_sheet(specs)}


def dims_from_stage1(s1):
    """stage1 → FAI dims [(中心,上,下)], 跳过①豁免的尺寸(断点2: 豁免尺寸不写FAI规格)。"""
    exempt = {e.get("序号") for e in (s1.get("exemptions") or [])}
    return [(d.get("中心"), d.get("上"), d.get("下"))
            for i, d in enumerate(s1.get("dimensions") or []) if i not in exempt]


def assemble_job(job, blank=BLANK):
    """从 .work/orders/<job> 现场读 stage3/stage1/photos → 装配终态承认书。供导出端点(子进程)调。"""
    from app import state                       # 延迟导入避免循环
    s3 = state.load_json(job, "stage3_filetree.json") or state.load_json(job, "stage2_bom.json", {})
    if not s3 or not s3.get("materials"):
        return {"ok": False, "err": "无 BOM/文件树数据(先完成③④)"}
    s1 = state.load_json(job, "stage1_drawing.json", {})
    meta = {"名称": s1.get("名称") or "", "品号": s1.get("品号", ""), "版本": s1.get("版本", "")}
    dims = dims_from_stage1(s1)
    photos = [os.path.join(state.photos_dir(job), p) for p in state.photos_list(job)]
    code = meta["品号"] or job
    outdir = os.path.join(ROOT, "产出留档", "导出", code)
    out = os.path.join(outdir, f"{code}_承认书.xlsx")
    try:
        r = assemble(s3, meta, dims, state.materials_dir(job), state.drawing_pdf(job),
                     out, outdir, photos, blank)
        r["ok"] = True
        return r
    except Exception as e:
        return {"ok": False, "err": str(e)[:200]}


if __name__ == "__main__":                      # 子进程入口: python -m hitl.assemble_order <job>
    import io
    import json
    import sys
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
    res = assemble_job(sys.argv[1]) if len(sys.argv) > 1 else {"ok": False, "err": "缺 job 参数"}
    print("@@RESULT@@" + json.dumps(res, ensure_ascii=False))
