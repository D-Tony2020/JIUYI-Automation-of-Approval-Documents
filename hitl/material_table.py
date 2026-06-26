# -*- coding: utf-8 -*-
"""W4 材质成分展开表：表头 + 变行数据区(LLM JOIN) + 浮动底部 RoHS 排除表。

- 抽取(mock/live) → assemble_row(归一, 含 weight 单位嗅探修复) → 嵌套合并注入(复用 m7 结构)。
- 底部 RoHS 排除表 = 固定静态块，从模板捕获后浮动摆放到数据区之后(签认: 浮动)。
- J2 填表日期: 清模板硬编码的 2026-06-18, 填生成日。
"""
import os
import re
import copy
import json
import datetime
from openpyxl.styles import Border, Side, Alignment, Font

# 成份名称标准词表(由7样本实证归纳): {材质: {cas: 标准名}}。查不到→保留MSDS原文。
_DICT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "成份名称词表.json")
try:
    with open(_DICT_PATH, encoding="utf-8") as _f:
        _COMP_DICT = json.load(_f)
except Exception:
    _COMP_DICT = {}


def normalize_component_name(材质, cas, raw):
    """按 (材质,CAS) 查标准词表归一成份名；查不到则保留 MSDS 原文(交人工)。"""
    std = _COMP_DICT.get((材质 or "").strip(), {}).get((cas or "").strip())
    return std if std else (raw or "").strip()

MAT_SHEET = "7.材质成分展开表 "

THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
FONT = Font(name="宋体", size=9)

ROHS_KEYS = ["Pb", "Cd", "Hg", "Cr6+", "PBBs", "PBDEs", "DEHP", "DBP", "BBP", "DIBP"]
ROHS_COL = {k: 13 + i for i, k in enumerate(ROHS_KEYS)}   # M..V
PART_MERGE_COLS = [1, 2, 3]                # A,B,C 跨整个零件
MAT_MERGE_COLS = [4, 5, 6]                 # D材质类别,E材质,F材质重量 跨整个材质
BLOCK_MERGE_COLS = [11, 12] + list(range(13, 31))  # K,L,M..AD 跨同一报告块(支持多色)
DATA_TOP = 14

SUPPLIER_ALIAS = {
    "兴鸿泰": ["兴鸿泰", "興鴻泰", "兴鸿泰锡业", "興鴻泰錫業", "深圳市兴鸿泰锡业有限公司", "XING HONG TAI"],
}
_MONTHS = {m: i + 1 for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"])}
_CAS_RE = re.compile(r"^\d{2,7}-\d{2}-\d$")


def _fmt(v):
    return ("%f" % v).rstrip("0").rstrip(".")


def normalize_cas(raw):
    if not raw:
        return ""
    s = re.sub(r"\s+", "", str(raw))
    return s if _CAS_RE.match(s) else str(raw).strip()


def normalize_supplier(raw):
    raw = (raw or "").strip()
    for canon, aliases in SUPPLIER_ALIAS.items():
        if any(a.lower() in raw.lower() for a in aliases):
            return canon
    return raw


def normalize_weight(raw):
    """单位嗅探(签认): 带%或>1→÷100; ≤1且无%→已是小数不除; <3/>x/余量 原样。"""
    if raw is None:
        return ""
    s = str(raw).strip()
    if "余量" in s or s.lower() in ("balance", "bal"):
        return "余量"
    if s and s[0] in "<≤>":
        return s.replace(" ", "")
    has_pct = "%" in s
    m = re.search(r"[-+]?\d*\.?\d+", s)
    if not m:
        return s
    val = float(m.group(0))
    if has_pct or val > 1:
        return _fmt(val / 100.0)
    return _fmt(val)   # ≤1 且无% → 已是小数, 不再除(needs_review 另行 track)


def normalize_date(raw):
    if not raw:
        return ""
    s = str(raw).strip()
    m = re.search(r"(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", s)
    if m:
        return f"{m.group(1)}.{int(m.group(2)):02d}.{int(m.group(3)):02d}"
    m = re.search(r"([A-Za-z]{3})[a-z]*\.?\s+(\d{1,2}),?\s+(\d{4})", s)
    if m:
        mon = _MONTHS.get(m.group(1).lower())
        if mon:
            return f"{m.group(3)}.{mon:02d}.{int(m.group(2)):02d}"
    return s


def normalize_rohs(result):
    s = str(result).strip()
    if s.upper().replace(".", "") == "ND":
        return "ND"
    m = re.search(r"[-+]?\d*\.?\d+", s)
    return m.group(0) if m else s


def assemble_row(msds, rohs, 零件, 材质类别, 材质):
    """MSDS 抽取 ⋈ RoHS 抽取 → 材质表一行(归一后)。"""
    成份 = []
    for c in msds.get("components", []):
        cas = normalize_cas(c.get("cas", ""))
        成份.append({
            "成份名称": normalize_component_name(材质, cas, c.get("name", "")),
            "CAS": cas,
            "重量%": normalize_weight(c.get("weight_pct_raw", "")),
        })
    rin = rohs.get("rohs", {})
    rout = {}
    for k in ROHS_KEYS:
        cell = rin.get(k, {})
        rout[k] = normalize_rohs(cell.get("result", "")) if isinstance(cell, dict) else normalize_rohs(cell)
    return {"零件": 零件, "材质类别": 材质类别, "材质": 材质,
            "供应商": normalize_supplier(msds.get("supplier_name_raw", "")),
            "成份": 成份, "RoHS": rout,
            "检测报告编号": (rohs.get("report_number", "") or "").strip(),
            "检测报告日期": normalize_date(rohs.get("report_date_raw", ""))}


def _style(ws, r, c, align=CENTER):
    cell = ws.cell(r, c)
    cell.border = BORDER
    cell.alignment = align
    cell.font = FONT


BOTTOM_COLS = 13   # 附表宽度 A:M


def _capture_bottom_block(ws):
    """清区前捕获底部 RoHS 排除表整块：每格(值+字体+填充+对齐+边框) + 合并 + 行高。

    旧版只捕文本→浮动后字体/合并/格式全丢；现完整捕获以保真还原。
    """
    start = None
    for r in range(40, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and "RoHS排除管制对象" in str(v):
            start = r
            break
    if start is None:
        return None
    end = start
    for r in range(start, ws.max_row + 1):
        if any(ws.cell(r, c).value not in (None, "") for c in range(1, BOTTOM_COLS + 1)):
            end = r
    rows = []
    for r in range(start, end + 1):
        cells = []
        for c in range(1, BOTTOM_COLS + 1):
            cell = ws.cell(r, c)
            cells.append((cell.value, copy.copy(cell.font), copy.copy(cell.fill),
                          copy.copy(cell.alignment), copy.copy(cell.border)))
        rows.append(cells)
    merges = [(mr.min_row - start, mr.min_col, mr.max_row - start, mr.max_col)
              for mr in ws.merged_cells.ranges
              if start <= mr.min_row <= end and mr.max_col <= BOTTOM_COLS]
    heights = [ws.row_dimensions[r].height for r in range(start, end + 1)]
    return {"rows": rows, "merges": merges, "heights": heights}


def _clear_region(ws, top, bottom):
    """清区：解合并 + 清值 + 重置样式(避免原附表位置残留空样式格)。"""
    from openpyxl.styles import PatternFill
    for mr in list(ws.merged_cells.ranges):
        min_c, min_r, max_c, max_r = mr.bounds
        if min_r >= top and max_r <= bottom:
            ws.unmerge_cells(str(mr))
    for row in ws.iter_rows(min_row=top, max_row=bottom, min_col=1, max_col=30):
        for cell in row:
            cell.value = None
            cell.font = Font()
            cell.fill = PatternFill()
            cell.border = Border()
            cell.alignment = Alignment()


# 报告类型 → 材质表里承载该报告 OLE 的列（实证7样本: 恒为 K/L/Y）
OLE_COL = {"MSDS": 11, "REACH": 12, "RoHS": 25}   # K=MSDS, L=Reach, Y=RoHS报告


def compute_layout(bom, start_row=DATA_TOP):
    """⭐结构搭建步骤(文本填入 + OLE装配 共用): 由 BOM 推导材质表单元格结构。

    返回 {parts:[{idx, first, last, part, materials:[{first,last,material,
          blocks:[{first,last,block}]}]}], start_row, data_end}。
    每零件/材质/报告块的行区间据成份数动态算出——不假设固定布局。
    """
    r = start_row
    parts = []
    for idx, part in enumerate(bom, 1):
        pf = r
        mats = []
        for mat in part["materials"]:
            mf = r
            blks = []
            for block in mat["blocks"]:
                bf = r
                r += max(1, len(block.get("成份", [])))
                blks.append({"first": bf, "last": r - 1, "block": block})
            mats.append({"first": mf, "last": r - 1, "material": mat, "blocks": blks})
        parts.append({"idx": idx, "first": pf, "last": r - 1, "part": part, "materials": mats})
    return {"parts": parts, "start_row": start_row, "data_end": r - 1}


def material_ole_anchors(bom, golden_ole, start_row=DATA_TOP):
    """⭐OLE装配也由结构驱动: golden 材质表OLE → 本单结构落位 [(row, col)]。

    golden 行 → 所属报告块 → 本单该块首行(compute_layout 算出)；列保持 K/L/Y。
    同(块,列)多个则块内顺延。换产品时块行随 BOM 变, OLE 位置自动跟着对。
    """
    layout = compute_layout(bom, start_row)
    blocks = [(B["first"], B["last"]) for P in layout["parts"]
              for M in P["materials"] for B in M["blocks"]]

    def block_first(grow):
        for f, l in blocks:
            if f <= grow <= l:
                return f
        return min(blocks, key=lambda fl: abs(fl[0] - grow))[0]

    used = {}
    res = []
    for g in golden_ole:
        bf = block_first(g["row"])
        key = (bf, g["col"])
        rr = bf + used.get(key, 0)
        used[key] = used.get(key, 0) + 1
        res.append((rr, g["col"]))
    return res


def inject_data(ws, bom, start_row=DATA_TOP):
    """按 compute_layout 的结构三层嵌套注入(零件A/B/C → 材质D/E/F → 报告块K..AD)。返回数据末行。"""
    layout = compute_layout(bom, start_row)
    for P in layout["parts"]:
        cat_runs = []      # 材质类别(D)按"零件内连续同类别"组合并: [[类别, first_row, last_row]]
        for M in P["materials"]:
            for B in M["blocks"]:
                block, bf, be = B["block"], B["first"], B["last"]
                材 = M["material"].get("材质", "")
                for j, comp in enumerate(block.get("成份", [])):
                    rr = bf + j
                    # G列成份名: 按(材质,CAS)查词表→承认书标准短名(B原文/MSDS全称→标准, 查不到留原文)
                    ws.cell(rr, 7, normalize_component_name(材, comp.get("CAS", ""), comp.get("成份名称", "")))
                    ws.cell(rr, 8, comp.get("CAS", ""))
                    ws.cell(rr, 10, comp.get("重量%", ""))
                for k in ROHS_KEYS:
                    ws.cell(bf, ROHS_COL[k], block.get("RoHS", {}).get(k, ""))
                ws.cell(bf, 23, block.get("报告日期", ""))
                ws.cell(bf, 24, block.get("报告编号", ""))
                ws.cell(bf, 9, "/")
                ws.cell(bf, 26, "/")
                ws.cell(bf, 27, "/")
                ws.cell(bf, 28, "Yes")
                ws.cell(bf, 29, "否")
                ws.cell(bf, 30, "/")
                if be > bf:
                    for c in BLOCK_MERGE_COLS:
                        ws.merge_cells(start_row=bf, start_column=c, end_row=be, end_column=c)
            mat, mf, me = M["material"], M["first"], M["last"]
            ws.cell(mf, 5, mat.get("材质", ""))      # E材质/F重量 按材质
            ws.cell(mf, 6, "/")
            if me > mf:
                for c in (5, 6):
                    ws.merge_cells(start_row=mf, start_column=c, end_row=me, end_column=c)
            cat = (mat.get("材质类别") or "").strip()    # 累计类别组(连续同类别合并 D)
            if cat_runs and cat_runs[-1][0] == cat:
                cat_runs[-1][2] = me
            else:
                cat_runs.append([cat, mf, me])
        for cat, cf, cl in cat_runs:                 # 写 D 材质类别 + 按组合并(一零件可多类别)
            ws.cell(cf, 4, cat)
            if cl > cf:
                ws.merge_cells(start_row=cf, start_column=4, end_row=cl, end_column=4)
        part, pf, pe = P["part"], P["first"], P["last"]
        ws.cell(pf, 1, P["idx"])
        ws.cell(pf, 2, part.get("零件", ""))
        ws.cell(pf, 3, part.get("供应商", ""))
        if pe > pf:
            for c in PART_MERGE_COLS:
                ws.merge_cells(start_row=pf, start_column=c, end_row=pe, end_column=c)
        for rr in range(pf, pe + 1):
            for cc in range(1, 31):
                _style(ws, rr, cc)
    return layout["data_end"]


def place_bottom_block(ws, block, start_row):
    """把捕获的底部排除表浮动摆到 start_row：完整还原值+字体+填充+对齐+边框+合并+行高。"""
    rows = block["rows"]
    for i, cells in enumerate(rows):
        r = start_row + i
        for c, (val, font, fill, align, border) in enumerate(cells, 1):
            cell = ws.cell(r, c)
            cell.value = val
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = border
        if block["heights"][i]:
            ws.row_dimensions[r].height = block["heights"][i]
    for dr1, c1, dr2, c2 in block["merges"]:
        ws.merge_cells(start_row=start_row + dr1, start_column=c1,
                       end_row=start_row + dr2, end_column=c2)
    return start_row + len(rows) - 1


def fill_material_table(ws, bom, product, gap=2):
    """表头 + 数据区 + 浮动底部表。返回 (数据末行, 底部表起始行)。"""
    block = _capture_bottom_block(ws)
    _clear_region(ws, DATA_TOP, ws.max_row)
    # 表头
    ws["A2"] = f"供货商类别  :{product['材料名称']}"
    today = product.get("填表日期") or datetime.date.today()
    if isinstance(today, datetime.date):
        ws["J2"] = f"填表日期 :                 {today.year} 年   {today.month:02d}   月  {today.day:02d}   日"
    else:
        ws["J2"] = f"填表日期 :                 {today}"
    # 数据区
    last = inject_data(ws, bom, DATA_TOP)
    # 浮动底部表
    bottom_start = last + 1 + gap
    if block:
        place_bottom_block(ws, block, bottom_start)
    return last, bottom_start


def selfcheck_material(ws, checks):
    """读回断言关键格。checks={coord: expected}。返回错误列表。"""
    errs = []
    for coord, exp in checks.items():
        got = ws[coord].value
        if str(got) != str(exp):
            errs.append(f"{coord}: 期望 {exp!r}，实得 {got!r}")
    return errs
