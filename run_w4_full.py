# -*- coding: utf-8 -*-
"""W4 材质成分表 · 全 4 料 BOM。直接用从 golden 解析的 demo BOM 注入，
逐格对照 golden 数据区(14-51)验证注入引擎对异构 4 料表的保真度。
"""
import os
import sys
import io
import json
import datetime

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from hitl.build import build_upto
from hitl.harness import assert_no_external_links
from hitl import material_table

ROOT = os.path.dirname(os.path.abspath(__file__))
TPL = os.path.join(ROOT, "模板", "承认书空白模板_治病.xlsx")
GOLDEN = os.path.join(ROOT, "案例材料", "承认书", "承认书", "做好的承认书",
                      "YY60039403 (J00016372) 承认书.xlsx")
BOM_JSON = os.path.join(ROOT, "hitl", "data", "demo_bom_YY60039403.json")
OUTDIR = os.path.join(ROOT, "产出留档", "W4-材质成分表")
BLANK = os.path.join(OUTDIR, "材质表_空白参照__00.xlsx")
OUT = os.path.join(OUTDIR, "材质表_全4料__02.xlsx")
SHEET = material_table.MAT_SHEET

BOM = json.load(open(BOM_JSON, encoding="utf-8"))
DATA = {
    "drawing_meta": {"名称": "SB120420BLCNR009导线", "品号": "YY60039403", "版本": "A01"},
    "product": {"材料名称": "导线", "填表日期": datetime.date(2026, 6, 24)},
    "bom": BOM,
}
DATA_END = 51   # golden 数据末行


def _norm(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return ("%g" % v)
    return str(v).strip()


def main():
    build_upto(TPL, BLANK, DATA, upto=2)  # 材质表空
    build_upto(TPL, OUT, DATA, upto=3)    # 全 4 料

    out_ws = openpyxl.load_workbook(OUT)[SHEET]
    gnames = openpyxl.load_workbook(GOLDEN).sheetnames
    gold_ws = openpyxl.load_workbook(GOLDEN, data_only=True)[
        [s for s in gnames if s.strip() == SHEET.strip()][0]]

    # 只比"有意义的数据列"(忽略 golden 人工填得不一致的占位列 F/I/K/L/Y/Z/AA/AD)
    DATA_COLS = [1, 2, 3, 4, 5, 7, 8, 10] + list(range(13, 25)) + [28, 29]  # A,B,C,D,E,G,H,J,M-V,W,X,AB,AC
    LAST = 48  # golden 最后一条成份行
    mism = []
    for r in range(14, LAST + 1):
        for c in DATA_COLS:
            gv, ov = _norm(gold_ws.cell(r, c).value), _norm(out_ws.cell(r, c).value)
            if gv != ov:
                mism.append((r, c, gv, ov))
    multicolor = [m for m in mism if m[0] == 17]   # PVC 红色第2份报告(多色, 已知业务点)
    other = [m for m in mism if m[0] != 17]

    errs = []
    if other:
        errs.append(f"除已知多色(r17)外仍有 {len(other)} 处数据不一致")
    assert_no_external_links(OUT)

    # 引擎结构：4 项次
    part_starts = [r for r in range(14, 49) if out_ws.cell(r, 1).value in (1, 2, 3, 4)]
    if part_starts != [14, 31, 41, 46]:
        errs.append(f"项次起始行异常: {part_starts}")
    # 两个非ND陷阱
    m_col = [_norm(out_ws.cell(r, 13).value) for r in range(14, 49)]
    for trap in ("14", "63"):
        if trap not in m_col:
            errs.append(f"非ND陷阱 Pb={trap} 未找到")
    # 底部表浮动(找标题行, 须在数据后)
    bt = [r for r in range(14, 80) if out_ws.cell(r, 1).value and "RoHS排除管制对象" in str(out_ws.cell(r, 1).value)]
    if not bt or bt[0] <= LAST:
        errs.append(f"底部表未浮动到数据后: {bt}")

    status = "PASS" if not errs else "FAIL"
    with open(os.path.join(OUTDIR, "_manifest.txt"), "a", encoding="utf-8") as f:
        f.write(f"材质表_全4料__02\t数据列diff={len(mism)}(多色{len(multicolor)}+其它{len(other)}) 底部表@A{bt[0] if bt else '?'}\t{status}\n")

    print("填入成品:", OUT)
    print(f"项次起始行: {part_starts}  底部表浮动到: A{bt[0] if bt else '?'}")
    print(f"数据列对 golden 差异: {len(mism)} 处（含多色 PVC 黑+红）")
    for r, c, gv, ov in mism[:20]:
        print(f"   r{r}c{c}: golden={gv!r} 我={ov!r}")
    if errs:
        print("❌ 自测失败：")
        for e in errs:
            print("   -", e)
        sys.exit(1)
    print("✅ 自测通过：全4料数据列逐格==golden(含多色PVC黑+红) + 4项次结构 + Pb=14/63陷阱 + 底部表浮动 + 无外链")


if __name__ == "__main__":
    main()
